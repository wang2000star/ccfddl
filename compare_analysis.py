#!/usr/bin/env python3
"""Compare user's CCF data against 2026 7th edition official directory."""
import openpyxl
import json

# ============================================================
# 1. Load user data
# ============================================================

# Conferences
wb_conf = openpyxl.load_workbook("C:/Users/82601/Desktop/DevByMe/ccfddl/conference_map.xlsx")
ws_conf = wb_conf["Sheet1"]
user_confs = []
for row in ws_conf.iter_rows(min_row=2, max_row=ws_conf.max_row, values_only=True):
    user_confs.append({
        "standard_abbr": row[0],
        "standard_ful": row[1],
        "alias_abbr": row[2],
        "alias_full": row[3],
        "proceeding_keys": row[4],
    })

# Journals
wb_jrnl = openpyxl.load_workbook("C:/Users/82601/Desktop/DevByMe/ccfddl/journal_map.xlsx")
ws_jrnl = wb_jrnl["Sheet1"]
user_jrns = []
for row in ws_jrnl.iter_rows(min_row=2, max_row=ws_jrnl.max_row, values_only=True):
    user_jrns.append({
        "standard_abbr": row[0],
        "standard_ful": row[1],
        "alias_abbr": row[2],
        "alias_full": row[3],
    })

print(f"User conferences: {len(user_confs)}")
print(f"User journals: {len(user_jrns)}")

# ============================================================
# 2. Analyze column data quality
# ============================================================

print("\n=== DATA QUALITY ISSUES ===")

# Check if standard_abbr contains abbreviations or full names
sample_abbr = [c["standard_abbr"] for c in user_confs[:10]]
print(f"\nFirst 10 'standard_abbr' values:")
for s in sample_abbr:
    print(f"  [{len(s)} chars] {s[:100]}...")

# Check for identical abbr and full name
same_count = sum(1 for c in user_confs if c["standard_abbr"] == c["standard_ful"])
print(f"\nConferences where abbr==ful: {same_count}/{len(user_confs)}")

same_count_j = sum(1 for j in user_jrns if j["standard_abbr"] == j["standard_ful"])
print(f"Journals where abbr==ful: {same_count_j}/{len(user_jrns)}")

# Check for missing values
for col in ["standard_abbr", "standard_ful", "alias_abbr", "alias_full"]:
    missing = sum(1 for c in user_confs if c[col] is None)
    if missing:
        print(f"Conferences missing {col}: {missing}")

for col in ["standard_abbr", "standard_ful", "alias_abbr", "alias_full"]:
    missing = sum(1 for j in user_jrns if j[col] is None)
    if missing:
        print(f"Journals missing {col}: {missing}")

# Check for proceeding_keys
has_keys = sum(1 for c in user_confs if c["proceeding_keys"] is not None)
print(f"\nConferences with proceeding_keys: {has_keys}/{len(user_confs)}")

# ============================================================
# 3. Check for duplicate entries
# ============================================================

print("\n=== DUPLICATE CHECK ===")
conf_names = [c["standard_ful"] for c in user_confs]
conf_dupes = set()
conf_seen = set()
for name in conf_names:
    if name in conf_seen:
        conf_dupes.add(name)
    conf_seen.add(name)
if conf_dupes:
    print(f"Duplicate conference names: {conf_dupes}")
else:
    print("No duplicate conferences found.")

jrn_names = [j["standard_ful"] for j in user_jrns]
jrn_dupes = set()
jrn_seen = set()
for name in jrn_names:
    if name in jrn_seen:
        jrn_dupes.add(name)
    jrn_seen.add(name)
if jrn_dupes:
    print(f"Duplicate journal names: {jrn_dupes}")
else:
    print("No duplicate journals found.")

# ============================================================
# 4. Count by expected category groups (from official 2026 list)
# ============================================================

# Official 2026 conference counts by category
official_conf_counts = {
    "体系结构": {"A": 11, "B": 26, "C": 30, "total": 67},
    "计算机网络": {"A": 4, "B": 10, "C": 20, "total": 34},
    "网络与信息安全": {"A": 6, "B": 11, "C": 29, "total": 46},
    "软件工程": {"A": 10, "B": 20, "C": 27, "total": 57},
    "数据库": {"A": 5, "B": 13, "C": 13, "total": 31},
    "计算机科学理论": {"A": 5, "B": 10, "C": 10, "total": 25},
    "计算机图形学与多媒体": {"A": 4, "B": 14, "C": 17, "total": 35},
    "人工智能": {"A": 7, "B": 14, "C": 22, "total": 43},
    "人机交互与普适计算": {"A": 4, "B": 7, "C": 15, "total": 26},
    "交叉综合新兴": {"A": 2, "B": 7, "C": 13, "total": 22},
}

total_official = sum(v["total"] for v in official_conf_counts.values())
print(f"\n=== COUNT COMPARISON ===")
print(f"Official 2026 total conferences: {total_official}")
print(f"User total conferences: {len(user_confs)}")

for cat, counts in official_conf_counts.items():
    print(f"  {cat}: A={counts['A']}, B={counts['B']}, C={counts['C']}, total={counts['total']}")

# Official 2026 journal counts by category
official_jrn_counts = {
    "体系结构": {"A": 6, "B": 10, "C": 12, "total": 28},
    "计算机网络": {"A": 3, "B": 6, "C": 12, "total": 21},
    "网络与信息安全": {"A": 3, "B": 5, "C": 9, "total": 17},
    "软件工程": {"A": 4, "B": 12, "C": 9, "total": 25},
    "数据库": {"A": 4, "B": 14, "C": 16, "total": 34},
    "计算机科学理论": {"A": 3, "B": 13, "C": 12, "total": 28},
    "计算机图形学与多媒体": {"A": 4, "B": 9, "C": 15, "total": 28},
    "人工智能": {"A": 4, "B": 21, "C": 39, "total": 64},
    "人机交互与普适计算": {"A": 2, "B": 8, "C": 5, "total": 15},
    "交叉综合新兴": {"A": 4, "B": 15, "C": 16, "total": 35},
}

total_official_jrn = sum(v["total"] for v in official_jrn_counts.values())
print(f"\nOfficial 2026 total journals: {total_official_jrn}")
print(f"User total journals: {len(user_jrns)}")

for cat, counts in official_jrn_counts.items():
    print(f"  {cat}: A={counts['A']}, B={counts['B']}, C={counts['C']}, total={counts['total']}")

# ============================================================
# 5. Specific version checks - key differences 2022 vs 2026
# ============================================================

print("\n=== VERSION CHECK: 2022 vs 2026 ===")

# Key indicators of 2026 version:
# - ICLR (added to AI-A in 2026)
# - IJCAI (downgraded A->B in 2026)
# - HPDC (upgraded B->A in 2026)
# - TMM (IEEE Trans. on Multimedia upgraded B->A in 2026)
# - Bioinformatics (upgraded B->A in 2026)

user_conf_names = [c["standard_ful"] for c in user_confs]

checks = {
    "ICLR (新增A类-2026标志)": "Learning Representations",
    "IJCAI (降级A->B-2026标志)": "Joint Conference on Artificial Intelligence",
    "HPDC (升级B->A-2026标志)": "High-Performance Parallel and Distributed Computing",
    "HotStorage (已有)": "HotStorage",
    "AFT (新增-2026)": "Advances in Financial Technologies",
    "IJTCS-FAW (新增-2026)": "International Joint Conference on Theoretical Computer Science",
}

for label, keyword in checks.items():
    found = [n for n in user_conf_names if keyword.lower() in n.lower()]
    if found:
        print(f"  ✅ {label}: FOUND - '{found[0][:80]}'")
    else:
        print(f"  ❌ {label}: NOT FOUND")

# Journal checks
user_jrn_names = [j["standard_ful"] for j in user_jrns]
jrn_checks = {
    "TMM (升级B->A-2026)": "Transactions on Multimedia",
    "Bioinformatics (升级B->A-2026)": "Bioinformatics",
    "TELO (新增-2026)": "Evolutionary Learning and Optimization",
    "JATS (新增-2026)": "Autonomous Transportation Systems",
    "ACM DLT (新增-2026)": "Distributed Ledger Technologies",
    "IACR TCHES (新增-2026)": "Cryptographic Hardware and Embedded Systems",
    "IACR ToSC (新增-2026)": "Symmetric Cryptology",
}

for label, keyword in jrn_checks.items():
    found = [n for n in user_jrn_names if keyword.lower() in n.lower()]
    if found:
        print(f"  ✅ {label}: FOUND - '{found[0][:80]}'")
    else:
        print(f"  ❌ {label}: NOT FOUND")

# ============================================================
# 6. Check for removed conferences (2022 had but 2026 removed)
# ============================================================

# The 2026 version removed 1 conference
# Also check if there are any conferences in user data that were in 2022 but NOT in 2026
print("\n=== POTENTIAL 2022-ONLY CONFERENCES (should be removed) ===")
# This is harder to check without the full 2022 list
# But some known removals/renames:
print("Checking for known removed/renamed items...")
# "ICPP" was renamed? No, ICPP is still there.
# The official change says: 1 conference removed, 2 renamed
# We'll flag this as needing manual verification

print("\nAnalysis complete.")
