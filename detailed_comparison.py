#!/usr/bin/env python3
"""Detailed comparison - find exactly which conferences/journals differ."""
import openpyxl
import re

# ============================================================
# Load user data
# ============================================================
wb_conf = openpyxl.load_workbook("C:/Users/82601/Desktop/DevByMe/ccfddl/conference_map.xlsx")
ws_conf = wb_conf["Sheet1"]
user_confs = []
for row in ws_conf.iter_rows(min_row=2, max_row=ws_conf.max_row, values_only=True):
    user_confs.append(row[1])  # standard_ful

wb_jrnl = openpyxl.load_workbook("C:/Users/82601/Desktop/DevByMe/ccfddl/journal_map.xlsx")
ws_jrnl = wb_jrnl["Sheet1"]
user_jrns = []
for row in ws_jrnl.iter_rows(min_row=2, max_row=ws_jrnl.max_row, values_only=True):
    user_jrns.append(row[1])

# ============================================================
# Official 2026 Conference list with FULL NAMES and RANKS
# ============================================================
official_conf_list = [
    # 体系结构 A (11)
    ("PPoPP", "ACM SIGPLAN Symposium on Principles & Practice of Parallel Programming", "A", "体系结构"),
    ("FAST", "USENIX Conference on File and Storage Technologies", "A", "体系结构"),
    ("DAC", "Design Automation Conference", "A", "体系结构"),
    ("HPCA", "IEEE International Symposium on High Performance Computer Architecture", "A", "体系结构"),
    ("MICRO", "IEEE/ACM International Symposium on Microarchitecture", "A", "体系结构"),
    ("SC", "International Conference for High Performance Computing, Networking, Storage, and Analysis", "A", "体系结构"),
    ("ASPLOS", "International Conference on Architectural Support for Programming Languages and Operating Systems", "A", "体系结构"),
    ("ISCA", "International Symposium on Computer Architecture", "A", "体系结构"),
    ("ATC", "USENIX Annual Technical Conference", "A", "体系结构"),
    ("EuroSys", "European Conference on Computer Systems", "A", "体系结构"),
    ("HPDC", "International ACM Symposium on High-Performance Parallel and Distributed Computing", "A", "体系结构"),
    # 体系结构 B (26)
    ("SoCC", "ACM Symposium on Cloud Computing", "B", "体系结构"),
    ("SPAA", "ACM Symposium on Parallelism in Algorithms and Architectures", "B", "体系结构"),
    ("PODC", "ACM Symposium on Principles of Distributed Computing", "B", "体系结构"),
    ("FPGA", "ACM/SIGDA International Symposium on Field-Programmable Gate Arrays", "B", "体系结构"),
    ("CGO", "International Symposium on Code Generation and Optimization", "B", "体系结构"),
    ("DATE", "Design, Automation & Test in Europe", "B", "体系结构"),
    ("HOT CHIPS", "Hot Chips: A Symposium on High Performance Chips", "B", "体系结构"),
    ("CLUSTER", "IEEE International Conference on Cluster Computing", "B", "体系结构"),
    ("ICCD", "International Conference on Computer Design", "B", "体系结构"),
    ("ICCAD", "International Conference on Computer-Aided Design", "B", "体系结构"),
    ("ICDCS", "IEEE International Conference on Distributed Computing Systems", "B", "体系结构"),
    ("CODES+ISSS", "International Conference on Hardware/Software Co-design and System Synthesis", "B", "体系结构"),
    ("HiPEAC", "International Conference on High Performance and Embedded Architectures and Compilers", "B", "体系结构"),
    ("SIGMETRICS", "International Conference on Measurement and Modeling of Computer Systems", "B", "体系结构"),
    ("PACT", "International Conference on Parallel Architectures and Compilation Techniques", "B", "体系结构"),
    ("ICPP", "International Conference on Parallel Processing", "B", "体系结构"),
    ("ICS", "International Conference on Supercomputing", "B", "体系结构"),
    ("VEE", "International Conference on Virtual Execution Environments", "B", "体系结构"),
    ("IPDPS", "IEEE International Parallel & Distributed Processing Symposium", "B", "体系结构"),
    ("Performance", "International Symposium on Computer Performance, Modeling, Measurements and Evaluation", "B", "体系结构"),
    ("ITC", "International Test Conference", "B", "体系结构"),
    ("LISA", "Large Installation System Administration Conference", "B", "体系结构"),
    ("MSST", "Mass Storage Systems and Technologies", "B", "体系结构"),
    ("RTAS", "IEEE Real-Time and Embedded Technology and Applications Symposium", "B", "体系结构"),
    ("Euro-Par", "European Conference on Parallel and Distributed Computing", "B", "体系结构"),
    ("ISCAS", "IEEE International Symposium on Circuits and Systems", "B", "体系结构"),
    # 体系结构 C (30)
    ("CF", "ACM International Conference on Computing Frontiers", "C", "体系结构"),
    ("SYSTOR", "ACM International Systems and Storage Conference", "C", "体系结构"),
    ("NOCS", "ACM/IEEE International Symposium on Networks-on-Chip", "C", "体系结构"),
    ("ASAP", "IEEE International Conference on Application-Specific Systems, Architectures, and Processors", "C", "体系结构"),
    ("ASP-DAC", "Asia and South Pacific Design Automation Conference", "C", "体系结构"),
    ("ETS", "IEEE European Test Symposium", "C", "体系结构"),
    ("FPL", "International Conference on Field-Programmable Logic and Applications", "C", "体系结构"),
    ("FCCM", "IEEE Symposium on Field-Programmable Custom Computing Machines", "C", "体系结构"),
    ("GLSVLSI", "Great Lakes Symposium on VLSI", "C", "体系结构"),
    ("ATS", "IEEE Asian Test Symposium", "C", "体系结构"),
    ("HPCC", "IEEE International Conference on High Performance Computing and Communications", "C", "体系结构"),
    ("HiPC", "IEEE International Conference on High Performance Computing, Data and Analytics", "C", "体系结构"),
    ("MASCOTS", "International Symposium on Modeling, Analysis, and Simulation of Computer and Telecommunication Systems", "C", "体系结构"),
    ("ISPA", "IEEE International Symposium on Parallel and Distributed Processing with Applications", "C", "体系结构"),
    ("CCGRID", "IEEE/ACM International Symposium on Cluster, Cloud and Grid Computing", "C", "体系结构"),
    ("NPC", "IFIP International Conference on Network and Parallel Computing", "C", "体系结构"),
    ("ICA3PP", "International Conference on Algorithms and Architectures for Parallel Processing", "C", "体系结构"),
    ("CASES", "International Conference on Compilers, Architectures, and Synthesis for Embedded Systems", "C", "体系结构"),
    ("FPT", "International Conference on Field-Programmable Technology", "C", "体系结构"),
    ("ICPADS", "International Conference on Parallel and Distributed Systems", "C", "体系结构"),
    ("ISLPED", "International Symposium on Low Power Electronics and Design", "C", "体系结构"),
    ("ISPD", "International Symposium on Physical Design", "C", "体系结构"),
    ("HOTI", "IEEE Symposium on High-Performance Interconnects", "C", "体系结构"),
    ("VTS", "IEEE VLSI Test Symposium", "C", "体系结构"),
    ("ITC-Asia", "International Test Conference in Asia", "C", "体系结构"),
    ("SEC", "ACM/IEEE Symposium on Edge Computing", "C", "体系结构"),
    ("NAS", "International Conference on Networking, Architecture and Storages", "C", "体系结构"),
    ("HotStorage", "HotStorage", "C", "体系结构"),
    ("APPT", "International Symposium on Advanced Parallel Processing Technology", "C", "体系结构"),
    ("JCC", "International Conference on JointCloud Computing", "C", "体系结构"),
    # 计算机网络 A (4)
    ("SIGCOMM", "ACM International Conference on Applications, Technologies, Architectures, and Protocols for Computer Communication", "A", "计算机网络"),
    ("MobiCom", "ACM International Conference on Mobile Computing and Networking", "A", "计算机网络"),
    ("INFOCOM", "IEEE International Conference on Computer Communications", "A", "计算机网络"),
    ("NSDI", "Symposium on Network System Design and Implementation", "A", "计算机网络"),
    # 计算机网络 B (10)
    ("SenSys", "ACM Conference on Embedded Networked Sensor Systems", "B", "计算机网络"),
    ("CoNEXT", "ACM International Conference on Emerging Networking Experiments and Technologies", "B", "计算机网络"),
    ("SECON", "IEEE International Conference on Sensing, Communication, and Networking", "B", "计算机网络"),
    ("IPSN", "International Conference on Information Processing in Sensor Networks", "B", "计算机网络"),
    ("MobiSys", "ACM International Conference on Mobile Systems, Applications, and Services", "B", "计算机网络"),
    ("ICNP", "IEEE International Conference on Network Protocols", "B", "计算机网络"),
    ("MobiHoc", "International Symposium on Theory, Algorithmic Foundations, and Protocol Design for Mobile Networks and Mobile Computing", "B", "计算机网络"),
    ("NOSSDAV", "International Workshop on Network and Operating System Support for Digital Audio and Video", "B", "计算机网络"),
    ("IWQoS", "IEEE/ACM International Workshop on Quality of Service", "B", "计算机网络"),
    ("IMC", "ACM Internet Measurement Conference", "B", "计算机网络"),
    # 计算机网络 C (20)
    ("ANCS", "ACM/IEEE Symposium on Architectures for Networking and Communication Systems", "C", "计算机网络"),
    ("APNOMS", "Asia-Pacific Network Operations and Management Symposium", "C", "计算机网络"),
    ("FORTE", "International Conference on Formal Techniques for Distributed Objects, Components, and Systems", "C", "计算机网络"),
    ("LCN", "IEEE Conference on Local Computer Networks", "C", "计算机网络"),
    ("GLOBECOM", "IEEE Global Communications Conference", "C", "计算机网络"),
    ("ICC", "IEEE International Conference on Communications", "C", "计算机网络"),
    ("ICCCN", "IEEE International Conference on Computer Communications and Networks", "C", "计算机网络"),
    ("MASS", "IEEE International Conference on Mobile Adhoc and Sensor Systems", "C", "计算机网络"),
    ("P2P", "IEEE International Conference on Peer-to-Peer Computing", "C", "计算机网络"),
    ("IPCCC", "IEEE International Performance Computing and Communications Conference", "C", "计算机网络"),
    ("WoWMoM", "IEEE International Symposium on a World of Wireless, Mobile and Multimedia Networks", "C", "计算机网络"),
    ("ISCC", "IEEE Symposium on Computers and Communications", "C", "计算机网络"),
    ("WCNC", "IEEE Wireless Communications and Networking Conference", "C", "计算机网络"),
    ("Networking", "IFIP International Conferences on Networking", "C", "计算机网络"),
    ("IM", "IFIP/IEEE International Symposium on Integrated Network Management", "C", "计算机网络"),
    ("MSN", "International Conference on Mobility, Sensing and Networking", "C", "计算机网络"),
    ("MSWiM", "International Conference on Modeling, Analysis and Simulation of Wireless and Mobile Systems", "C", "计算机网络"),
    ("WASA", "International Conference on Wireless Artificial Intelligent Computing Systems and Applications", "C", "计算机网络"),
    ("HotNets", "ACM Workshop on Hot Topics in Networks", "C", "计算机网络"),
    ("APNet", "Asia-Pacific Workshop on Networking", "C", "计算机网络"),
    # 信息安全 A (6)
    ("CCS", "ACM Conference on Computer and Communications Security", "A", "信息安全"),
    ("EUROCRYPT", "International Conference on the Theory and Applications of Cryptographic Techniques", "A", "信息安全"),
    ("S&P", "IEEE Symposium on Security and Privacy", "A", "信息安全"),
    ("CRYPTO", "International Cryptology Conference", "A", "信息安全"),
    ("USENIX Security", "USENIX Security Symposium", "A", "信息安全"),
    ("NDSS", "Network and Distributed System Security Symposium", "A", "信息安全"),
    # 信息安全 B (11)
    ("ACSAC", "Annual Computer Security Applications Conference", "B", "信息安全"),
    ("ASIACRYPT", "Annual International Conference on the Theory and Application of Cryptology and Information Security", "B", "信息安全"),
    ("ESORICS", "European Symposium on Research in Computer Security", "B", "信息安全"),
    ("FSE", "Fast Software Encryption", "B", "信息安全"),
    ("CSFW", "IEEE Computer Security Foundations Workshop", "B", "信息安全"),
    ("SRDS", "IEEE International Symposium on Reliable Distributed Systems", "B", "信息安全"),
    ("CHES", "International Conference on Cryptographic Hardware and Embedded Systems", "B", "信息安全"),
    ("DSN", "International Conference on Dependable Systems and Networks", "B", "信息安全"),
    ("RAID", "International Symposium on Recent Advances in Intrusion Detection", "B", "信息安全"),
    ("PKC", "International Workshop on Practice and Theory in Public Key Cryptography", "B", "信息安全"),
    ("TCC", "Theory of Cryptography Conference", "B", "信息安全"),
    # 信息安全 C (29)
    ("WiSec", "ACM Conference on Security and Privacy in Wireless and Mobile Networks", "C", "信息安全"),
    ("SACMAT", "ACM Symposium on Access Control Models and Technologies", "C", "信息安全"),
    ("DRM", "ACM Workshop on Digital Rights Management", "C", "信息安全"),
    ("IH&MMSec", "ACM Workshop on Information Hiding and Multimedia Security", "C", "信息安全"),
    ("ACNS", "International Conference on Applied Cryptography and Network Security", "C", "信息安全"),
    ("AsiaCCS", "ACM Asia Conference on Computer and Communications Security", "C", "信息安全"),
    ("ACISP", "Australasia Conference on Information Security and Privacy", "C", "信息安全"),
    ("CT-RSA", "The Cryptographer's Track at RSA Conference", "C", "信息安全"),
    ("DIMVA", "Conference on Detection of Intrusions and Malware & Vulnerability Assessment", "C", "信息安全"),
    ("DFRWS", "Digital Forensic Research Workshop", "C", "信息安全"),
    ("FC", "Financial Cryptography and Data Security", "C", "信息安全"),
    ("TrustCom", "IEEE International Conference on Trust, Security and Privacy in Computing and Communications", "C", "信息安全"),
    ("IFIP SEC", "IFIP International Information Security Conference", "C", "信息安全"),
    ("IFIP WG 11.9", "IFIP Working Group 11.9 International Conference on Digital Forensics", "C", "信息安全"),
    ("ISC", "Information Security Conference", "C", "信息安全"),
    ("ICDF2C", "International Conference on Digital Forensics & Cyber Crime", "C", "信息安全"),
    ("ICICS", "International Conference on Information and Communications Security", "C", "信息安全"),
    ("SecureComm", "International Conference on Security and Privacy in Communication Networks", "C", "信息安全"),
    ("NSPW", "New Security Paradigms Workshop", "C", "信息安全"),
    ("PAM", "Passive and Active Measurement Conference", "C", "信息安全"),
    ("PETS", "Privacy Enhancing Technologies Symposium", "C", "信息安全"),
    ("SAC", "Selected Areas in Cryptography", "C", "信息安全"),
    ("SOUPS", "Symposium On Usable Privacy and Security", "C", "信息安全"),
    ("HotSec", "USENIX Workshop on Hot Topics in Security", "C", "信息安全"),
    ("EuroS&P", "IEEE European Symposium on Security and Privacy", "C", "信息安全"),
    ("Inscrypt", "International Conference on Information Security and Cryptology", "C", "信息安全"),
    ("CODASPY", "Conference on Data and Application Security and Privacy", "C", "信息安全"),
    ("BlockSys", "International Conference on Blockchain, Artificial Intelligence, and Trustworthy Systems", "C", "信息安全"),
    ("CSCloud", "International Conference on Cyber Security and Cloud Computing", "C", "信息安全"),
    # 软件工程 A (10)
    ("PLDI", "ACM SIGPLAN Conference on Programming Language Design and Implementation", "A", "软件工程"),
    ("POPL", "ACM SIGPLAN-SIGACT Symposium on Principles of Programming Languages", "A", "软件工程"),
    ("FSE", "ACM International Conference on the Foundations of Software Engineering", "A", "软件工程"),
    ("SOSP", "ACM Symposium on Operating Systems Principles", "A", "软件工程"),
    ("OOPSLA", "Conference on Object-Oriented Programming Systems, Languages, and Applications", "A", "软件工程"),
    ("ASE", "International Conference on Automated Software Engineering", "A", "软件工程"),
    ("ICSE", "International Conference on Software Engineering", "A", "软件工程"),
    ("ISSTA", "International Symposium on Software Testing and Analysis", "A", "软件工程"),
    ("OSDI", "USENIX Symposium on Operating Systems Design and Implementation", "A", "软件工程"),
    ("FM", "International Symposium on Formal Methods", "A", "软件工程"),
    # 软件工程 B (20)
    ("ECOOP", "European Conference on Object-Oriented Programming", "B", "软件工程"),
    ("ETAPS", "European Joint Conferences on Theory and Practice of Software", "B", "软件工程"),
    ("ICPC", "IEEE International Conference on Program Comprehension", "B", "软件工程"),
    ("RE", "IEEE International Requirements Engineering Conference", "B", "软件工程"),
    ("CAiSE", "International Conference on Advanced Information Systems Engineering", "B", "软件工程"),
    ("ICFP", "ACM SIGPLAN International Conference on Function Programming", "B", "软件工程"),
    ("LCTES", "ACM SIGPLAN/SIGBED International Conference on Languages, Compilers and Tools for Embedded Systems", "B", "软件工程"),
    ("MoDELS", "ACM/IEEE International Conference on Model Driven Engineering Languages and Systems", "B", "软件工程"),
    ("CP", "International Conference on Principles and Practice of Constraint Programming", "B", "软件工程"),
    ("ICSOC", "International Conference on Service Oriented Computing", "B", "软件工程"),
    ("SANER", "IEEE International Conference on Software Analysis, Evolution, and Reengineering", "B", "软件工程"),
    ("ICSME", "International Conference on Software Maintenance and Evolution", "B", "软件工程"),
    ("VMCAI", "International Conference on Verification, Model Checking and Abstract Interpretation", "B", "软件工程"),
    ("ICWS", "IEEE International Conference on Web Services", "B", "软件工程"),
    ("Middleware", "International Middleware Conference", "B", "软件工程"),
    ("SAS", "International Static Analysis Symposium", "B", "软件工程"),
    ("ESEM", "International Symposium on Empirical Software Engineering and Measurement", "B", "软件工程"),
    ("ISSRE", "IEEE International Symposium on Software Reliability Engineering", "B", "软件工程"),
    ("HotOS", "USENIX Workshop on Hot Topics in Operating Systems", "B", "软件工程"),
    ("CC", "International Conference on Compiler Construction", "B", "软件工程"),
    # 软件工程 C (27)
    ("PEPM", "ACM SIGPLAN Workshop on Partial Evaluation and Program Manipulation", "C", "软件工程"),
    ("PASTE", "ACM SIGPLAN-SIGSOFT Workshop on Program Analysis for Software Tools and Engineering", "C", "软件工程"),
    ("APLAS", "Asian Symposium on Programming Languages and Systems", "C", "软件工程"),
    ("APSEC", "Asia-Pacific Software Engineering Conference", "C", "软件工程"),
    ("EASE", "International Conference on Evaluation and Assessment in Software Engineering", "C", "软件工程"),
    ("ICECCS", "International Conference on Engineering of Complex Computer Systems", "C", "软件工程"),
    ("ICST", "IEEE International Conference on Software Testing, Verification and Validation", "C", "软件工程"),
    ("ISPASS", "IEEE International Symposium on Performance Analysis of Systems and Software", "C", "软件工程"),
    ("SCAM", "IEEE International Working Conference on Source Code Analysis and Manipulation", "C", "软件工程"),
    ("COMPSAC", "International Computer Software and Applications Conference", "C", "软件工程"),
    ("ICFEM", "International Conference on Formal Engineering Methods", "C", "软件工程"),
    ("SSE", "IEEE International Conference on Software Services Engineering", "C", "软件工程"),
    ("ICSSP", "International Conference on Software and System Process", "C", "软件工程"),
    ("SEKE", "International Conference on Software Engineering and Knowledge Engineering", "C", "软件工程"),
    ("QRS", "International Conference on Software Quality, Reliability and Security", "C", "软件工程"),
    ("ICSR", "International Conference on Software Reuse", "C", "软件工程"),
    ("ICWE", "International Conference on Web Engineering", "C", "软件工程"),
    ("SPIN", "International Symposium on Model Checking of Software", "C", "软件工程"),
    ("ATVA", "International Symposium on Automated Technology for Verification and Analysis", "C", "软件工程"),
    ("LOPSTR", "International Symposium on Logic-based Program Synthesis and Transformation", "C", "软件工程"),
    ("TASE", "Theoretical Aspects of Software Engineering Conference", "C", "软件工程"),
    ("MSR", "Mining Software Repositories", "C", "软件工程"),
    ("REFSQ", "Requirements Engineering: Foundation for Software Quality", "C", "软件工程"),
    ("WICSA", "Working IEEE/IFIP Conference on Software Architecture", "C", "软件工程"),
    ("Internetware", "Asia-Pacific Symposium on Internetware", "C", "软件工程"),
    ("RV", "International Conference on Runtime Verification", "C", "软件工程"),
    ("MEMOCODE", "International Conference on Formal Methods and Models for Co-Design", "C", "软件工程"),
    # 数据库 A (5)
    ("SIGMOD", "ACM SIGMOD Conference", "A", "数据库"),
    ("SIGKDD", "ACM SIGKDD Conference on Knowledge Discovery and Data Mining", "A", "数据库"),
    ("ICDE", "IEEE International Conference on Data Engineering", "A", "数据库"),
    ("SIGIR", "International ACM SIGIR Conference on Research and Development in Information Retrieval", "A", "数据库"),
    ("VLDB", "International Conference on Very Large Data Bases", "A", "数据库"),
    # 数据库 B (13)
    ("CIKM", "ACM International Conference on Information and Knowledge Management", "B", "数据库"),
    ("WSDM", "ACM International Conference on Web Search and Data Mining", "B", "数据库"),
    ("PODS", "ACM SIGMOD-SIGACT-SIGAI Symposium on Principles of Database Systems", "B", "数据库"),
    ("DASFAA", "International Conference on Database Systems for Advanced Applications", "B", "数据库"),
    ("ECML-PKDD", "European Conference on Machine Learning and Principles and Practice of Knowledge Discovery in Databases", "B", "数据库"),
    ("ISWC", "IEEE International Semantic Web Conference", "B", "数据库"),
    ("ICDM", "IEEE International Conference on Data Mining", "B", "数据库"),
    ("ICDT", "International Conference on Database Theory", "B", "数据库"),
    ("EDBT", "International Conference on Extending Database Technology", "B", "数据库"),
    ("CIDR", "Conference on Innovative Data Systems Research", "B", "数据库"),
    ("SDM", "SIAM International Conference on Data Mining", "B", "数据库"),
    ("RecSys", "ACM Conference on Recommender Systems", "B", "数据库"),
    ("WISE", "Web Information Systems Engineering Conference", "B", "数据库"),
    # 数据库 C (13)
    ("APWeb", "Asia Pacific Web Conference", "C", "数据库"),
    ("DEXA", "International Conference on Database and Expert System Applications", "C", "数据库"),
    ("ECIR", "European Conference on Information Retrieval", "C", "数据库"),
    ("ESWC", "Extended Semantic Web Conference", "C", "数据库"),
    ("WebDB", "International Workshop on Web and Databases", "C", "数据库"),
    ("ER", "International Conference on Conceptual Modeling", "C", "数据库"),
    ("MDM", "International Conference on Mobile Data Management", "C", "数据库"),
    ("SSDBM", "International Conference on Scientific and Statistical Database Management", "C", "数据库"),
    ("WAIM", "International Conference on Web Age Information Management", "C", "数据库"),
    ("SSTD", "International Symposium on Spatial and Temporal Databases", "C", "数据库"),
    ("PAKDD", "Pacific-Asia Conference on Knowledge Discovery and Data Mining", "C", "数据库"),
    ("ADMA", "International Conference on Advanced Data Mining and Applications", "C", "数据库"),
    ("WISA", "Web Information Systems and Applications", "C", "数据库"),
    # 计算机科学理论 A (5)
    ("STOC", "ACM Symposium on the Theory of Computing", "A", "理论"),
    ("SODA", "ACM-SIAM Symposium on Discrete Algorithms", "A", "理论"),
    ("CAV", "International Conference on Computer Aided Verification", "A", "理论"),
    ("FOCS", "IEEE Annual Symposium on Foundations of Computer Science", "A", "理论"),
    ("LICS", "ACM/IEEE Symposium on Logic in Computer Science", "A", "理论"),
    # 理论 B (10)
    ("SoCG", "International Symposium on Computational Geometry", "B", "理论"),
    ("ESA", "European Symposium on Algorithms", "B", "理论"),
    ("CCC", "Conference on Computational Complexity", "B", "理论"),
    ("ICALP", "International Colloquium on Automata, Languages and Programming", "B", "理论"),
    ("CADE", "Conference on Automated Deduction", "B", "理论"),
    ("CONCUR", "International Conference on Concurrency Theory", "B", "理论"),
    ("HSCC", "International Conference on Hybrid Systems: Computation and Control", "B", "理论"),
    ("SAT", "International Conference on Theory and Applications of Satisfiability Testing", "B", "理论"),
    ("COCOON", "International Computing and Combinatorics Conference", "B", "理论"),
    ("FMCAD", "Formal Methods in Computer-Aided Design", "B", "理论"),
    # 理论 C (10)
    ("CSL", "Computer Science Logic", "C", "理论"),
    ("FSTTCS", "Foundations of Software Technology and Theoretical Computer Science", "C", "理论"),
    ("DSAA", "IEEE International Conference on Data Science and Advanced Analytics", "C", "理论"),
    ("ICTAC", "International Colloquium on Theoretical Aspects of Computing", "C", "理论"),
    ("IPCO", "International Conference on Integer Programming and Combinatorial Optimization", "C", "理论"),
    ("FSCD", "International Conference on Formal Structures for Computation and Deduction", "C", "理论"),
    ("ISAAC", "International Symposium on Algorithms and Computation", "C", "理论"),
    ("MFCS", "International Conference on Mathematical Foundations of Computer Science", "C", "理论"),
    ("STACS", "Symposium on Theoretical Aspects of Computer Science", "C", "理论"),
    ("SETTA", "International Symposium on Software Engineering: Theories, Tools, and Applications", "C", "理论"),
    # 图形学多媒体 A (4)
    ("ACM MM", "ACM International Conference on Multimedia", "A", "图形学多媒体"),
    ("SIGGRAPH", "ACM Special Interest Group on Computer Graphics", "A", "图形学多媒体"),
    ("VR", "IEEE Virtual Reality", "A", "图形学多媒体"),
    ("IEEE VIS", "IEEE Visualization Conference", "A", "图形学多媒体"),
    # 图形学多媒体 B (14)
    ("ICMR", "ACM SIGMM International Conference on Multimedia Retrieval", "B", "图形学多媒体"),
    ("I3D", "ACM SIGGRAPH Symposium on Interactive 3D Graphics and Games", "B", "图形学多媒体"),
    ("SCA", "ACM SIGGRAPH/Eurographics Symposium on Computer Animation", "B", "图形学多媒体"),
    ("DCC", "Data Compression Conference", "B", "图形学多媒体"),
    ("Eurographics", "Annual Conference of the European Association for Computer Graphics", "B", "图形学多媒体"),
    ("EuroVis", "Eurographics Conference on Visualization", "B", "图形学多媒体"),
    ("SGP", "Eurographics Symposium on Geometry Processing", "B", "图形学多媒体"),
    ("EGSR", "Eurographics Symposium on Rendering", "B", "图形学多媒体"),
    ("ICASSP", "IEEE International Conference on Acoustics, Speech and Signal Processing", "B", "图形学多媒体"),
    ("ICME", "IEEE International Conference on Multimedia & Expo", "B", "图形学多媒体"),
    ("ISMAR", "International Symposium on Mixed and Augmented Reality", "B", "图形学多媒体"),
    ("PG", "Pacific Conference on Computer Graphics and Applications", "B", "图形学多媒体"),
    ("SPM", "Symposium on Solid and Physical Modeling", "B", "图形学多媒体"),
    ("INTERSPEECH", "Conference of the International Speech Communication Association", "B", "图形学多媒体"),
    # 图形学多媒体 C (17)
    ("VRST", "ACM Symposium on Virtual Reality Software and Technology", "C", "图形学多媒体"),
    ("CASA", "International Conference on Computer Animation and Social Agents", "C", "图形学多媒体"),
    ("CGI", "Computer Graphics International", "C", "图形学多媒体"),
    ("GMP", "Geometric Modeling and Processing", "C", "图形学多媒体"),
    ("PacificVis", "IEEE Pacific Visualization Symposium", "C", "图形学多媒体"),
    ("3DV", "International Conference on 3D Vision", "C", "图形学多媒体"),
    ("CAD/Graphics", "International Conference on Computer-Aided Design and Computer Graphics", "C", "图形学多媒体"),
    ("ICIP", "IEEE International Conference on Image Processing", "C", "图形学多媒体"),
    ("MMM", "International Conference on Multimedia Modeling", "C", "图形学多媒体"),
    ("MMAsia", "ACM Multimedia Asia", "C", "图形学多媒体"),
    ("SMI", "Shape Modeling International", "C", "图形学多媒体"),
    ("CVM", "Computational Visual Media", "C", "图形学多媒体"),
    ("PRCV", "Chinese Conference on Pattern Recognition and Computer Vision", "C", "图形学多媒体"),
    ("ICIG", "International Conference on Image and Graphics", "C", "图形学多媒体"),
    ("NCMMSC", "National Conference on Man-Machine Speech Communication", "C", "图形学多媒体"),
    ("ASRU", "Automatic Speech Recognition and Understanding Workshop", "C", "图形学多媒体"),
    ("SLT", "Spoken Language Technology", "C", "图形学多媒体"),
    # 人工智能 A (7)
    ("AAAI", "AAAI Conference on Artificial Intelligence", "A", "人工智能"),
    ("NeurIPS", "Conference on Neural Information Processing Systems", "A", "人工智能"),
    ("ACL", "Annual Meeting of the Association for Computational Linguistics", "A", "人工智能"),
    ("CVPR", "IEEE/CVF Computer Vision and Pattern Recognition Conference", "A", "人工智能"),
    ("ICCV", "International Conference on Computer Vision", "A", "人工智能"),
    ("ICML", "International Conference on Machine Learning", "A", "人工智能"),
    ("ICLR", "International Conference on Learning Representations", "A", "人工智能"),
    # 人工智能 B (14)
    ("COLT", "Annual Conference on Computational Learning Theory", "B", "人工智能"),
    ("EMNLP", "Conference on Empirical Methods in Natural Language Processing", "B", "人工智能"),
    ("ECAI", "European Conference on Artificial Intelligence", "B", "人工智能"),
    ("ECCV", "European Conference on Computer Vision", "B", "人工智能"),
    ("ICRA", "IEEE International Conference on Robotics and Automation", "B", "人工智能"),
    ("ICAPS", "International Conference on Automated Planning and Scheduling", "B", "人工智能"),
    ("ICCBR", "International Conference on Case-Based Reasoning", "B", "人工智能"),
    ("COLING", "International Conference on Computational Linguistics", "B", "人工智能"),
    ("KR", "International Conference on Principles of Knowledge Representation and Reasoning", "B", "人工智能"),
    ("UAI", "Conference on Uncertainty in Artificial Intelligence", "B", "人工智能"),
    ("AAMAS", "International Joint Conference on Autonomous Agents and Multi-agent Systems", "B", "人工智能"),
    ("PPSN", "Parallel Problem Solving from Nature", "B", "人工智能"),
    ("NAACL", "North American Chapter of the Association for Computational Linguistics", "B", "人工智能"),
    ("IJCAI", "International Joint Conference on Artificial Intelligence", "B", "人工智能"),
    # 人工智能 C (22)
    ("AISTATS", "International Conference on Artificial Intelligence and Statistics", "C", "人工智能"),
    ("ACCV", "Asian Conference on Computer Vision", "C", "人工智能"),
    ("ACML", "Asian Conference on Machine Learning", "C", "人工智能"),
    ("BMVC", "British Machine Vision Conference", "C", "人工智能"),
    ("NLPCC", "CCF International Conference on Natural Language Processing and Chinese Computing", "C", "人工智能"),
    ("CoNLL", "Conference on Computational Natural Language Learning", "C", "人工智能"),
    ("GECCO", "Genetic and Evolutionary Computation Conference", "C", "人工智能"),
    ("ICTAI", "IEEE International Conference on Tools with Artificial Intelligence", "C", "人工智能"),
    ("IROS", "IEEE/RSJ International Conference on Intelligent Robots and Systems", "C", "人工智能"),
    ("ALT", "International Conference on Algorithmic Learning Theory", "C", "人工智能"),
    ("ICANN", "International Conference on Artificial Neural Networks", "C", "人工智能"),
    ("FG", "IEEE International Conference on Automatic Face and Gesture Recognition", "C", "人工智能"),
    ("ICDAR", "International Conference on Document Analysis and Recognition", "C", "人工智能"),
    ("ILP", "International Conference on Inductive Logic Programming", "C", "人工智能"),
    ("KSEM", "International Conference on Knowledge Science, Engineering and Management", "C", "人工智能"),
    ("ICONIP", "International Conference on Neural Information Processing", "C", "人工智能"),
    ("ICPR", "International Conference on Pattern Recognition", "C", "人工智能"),
    ("IJCB", "International Joint Conference on Biometrics", "C", "人工智能"),
    ("IJCNN", "International Joint Conference on Neural Networks", "C", "人工智能"),
    ("PRICAI", "Pacific Rim International Conference on Artificial Intelligence", "C", "人工智能"),
    ("IEEE CEC", "Congress on Evolutionary Computation", "C", "人工智能"),
    ("DAI", "International Conference on Distributed Artificial Intelligence", "C", "人工智能"),
    # 人机交互 A (4)
    ("CSCW", "ACM Conference On Computer-Supported Cooperative Work And Social Computing", "A", "人机交互"),
    ("CHI", "ACM Conference on Human Factors in Computing Systems", "A", "人机交互"),
    ("UbiComp", "ACM International Joint Conference on Pervasive and Ubiquitous Computing", "A", "人机交互"),
    ("UIST", "ACM Symposium on User Interface Software and Technology", "A", "人机交互"),
    # 人机交互 B (7)
    ("GROUP", "ACM International Conference on Supporting Group Work", "B", "人机交互"),
    ("IUI", "ACM International Conference on Intelligent User Interfaces", "B", "人机交互"),
    ("ISS", "ACM International Conference on Interactive Surfaces and Spaces", "B", "人机交互"),
    ("ECSCW", "European Conference on Computer Supported Cooperative Work", "B", "人机交互"),
    ("PERCOM", "IEEE International Conference on Pervasive Computing and Communications", "B", "人机交互"),
    ("MobileHCI", "ACM International Conference on Mobile Human-Computer Interaction", "B", "人机交互"),
    ("ICWSM", "International AAAI Conference on Web and Social Media", "B", "人机交互"),
    # 人机交互 C (15)
    ("DIS", "ACM SIGCHI Conference on Designing Interactive Systems", "C", "人机交互"),
    ("ICMI", "ACM International Conference on Multimodal Interaction", "C", "人机交互"),
    ("ASSETS", "International ACM SIGACCESS Conference on Computers and Accessibility", "C", "人机交互"),
    ("GI", "Graphics Interface", "C", "人机交互"),
    ("UIC", "IEEE International Conference on Ubiquitous Intelligence and Computing", "C", "人机交互"),
    ("World Haptics", "IEEE World Haptics Conference", "C", "人机交互"),
    ("INTERACT", "International Conference on Human-Computer Interaction of International Federation for Information Processing", "C", "人机交互"),
    ("IDC", "ACM Interaction Design and Children", "C", "人机交互"),
    ("CollaborateCom", "International Conference on Collaborative Computing: Networking, Applications and Worksharing", "C", "人机交互"),
    ("CSCWD", "International Conference on Computer Supported Cooperative Work in Design", "C", "人机交互"),
    ("CoopIS", "International Conference on Cooperative Information Systems", "C", "人机交互"),
    ("MobiQuitous", "International Conference on Mobile and Ubiquitous Systems: Computing, Networking and Services", "C", "人机交互"),
    ("AVI", "International Working Conference on Advanced Visual Interfaces", "C", "人机交互"),
    ("GPC", "Conference on Green, Pervasive and Cloud Computing", "C", "人机交互"),
    ("ICXR", "CCF International Conference on Extended Reality", "C", "人机交互"),
    # 交叉综合 A (2)
    ("WWW", "International World Wide Web Conference", "A", "交叉综合"),
    ("RTSS", "IEEE Real-Time Systems Symposium", "A", "交叉综合"),
    # 交叉综合 B (7)
    ("CogSci", "Annual Meeting of the Cognitive Science Society", "B", "交叉综合"),
    ("BIBM", "IEEE International Conference on Bioinformatics and Biomedicine", "B", "交叉综合"),
    ("EMSOFT", "International Conference on Embedded Software", "B", "交叉综合"),
    ("ISMB", "International Conference on Intelligent Systems for Molecular Biology", "B", "交叉综合"),
    ("RECOMB", "Annual International Conference on Research in Computational Molecular Biology", "B", "交叉综合"),
    ("MICCAI", "International Conference on Medical Image Computing and Computer-Assisted Intervention", "B", "交叉综合"),
    ("WINE", "Conference on Web and Internet Economics", "B", "交叉综合"),
    # 交叉综合 C (13)
    ("AMIA", "American Medical Informatics Association Annual Symposium", "C", "交叉综合"),
    ("APBC", "Asia Pacific Bioinformatics Conference", "C", "交叉综合"),
    ("IEEE BigData", "IEEE International Conference on Big Data", "C", "交叉综合"),
    ("IEEE CLOUD", "IEEE International Conference on Cloud Computing", "C", "交叉综合"),
    ("SMC", "IEEE International Conference on Systems, Man, and Cybernetics", "C", "交叉综合"),
    ("COSIT", "International Conference on Spatial Information Theory", "C", "交叉综合"),
    ("ISBRA", "International Symposium on Bioinformatics Research and Applications", "C", "交叉综合"),
    ("SAGT", "International Symposium on Algorithmic Game Theory", "C", "交叉综合"),
    ("SIGSPATIAL", "ACM Special Interest Group on Spatial Information", "C", "交叉综合"),
    ("ICIC", "International Conference on Intelligent Computing", "C", "交叉综合"),
    ("ICSS", "International Conference on Service Science", "C", "交叉综合"),
    ("AFT", "Advances in Financial Technologies", "C", "交叉综合"),
    ("IJTCS-FAW", "International Joint Conference on Theoretical Computer Science - Frontier of Algorithmic Wisdom", "C", "交叉综合"),
]

# Now do fuzzy matching
import difflib

def normalize(s):
    """Normalize string for comparison."""
    if s is None:
        return ""
    s = s.lower().strip()
    s = s.replace('\n', ' ').replace('\xa0', ' ')
    s = re.sub(r'\s+', ' ', s)
    s = s.replace('&', 'and')
    s = s.replace('  ', ' ')
    return s

def find_best_match(name, candidates, threshold=0.85):
    """Find best fuzzy match for a name in candidates."""
    name_norm = normalize(name)
    best_score = 0
    best_match = None
    for abbr, full, rank, cat in candidates:
        full_norm = normalize(full)
        score = difflib.SequenceMatcher(None, name_norm, full_norm).ratio()
        if score > best_score:
            best_score = score
            best_match = (abbr, full, rank, cat)
    if best_score >= threshold:
        return best_match, best_score
    return None, best_score

print("=" * 80)
print("CONFERENCE ANALYSIS")
print("=" * 80)

# Match each user conference to official list
matched = []
unmatched_user = []
for uc in user_confs:
    match, score = find_best_match(uc, official_conf_list, threshold=0.85)
    if match:
        matched.append((uc, match, score))
    else:
        unmatched_user.append((uc, score))

print(f"\nMatched: {len(matched)}")
print(f"Unmatched (user items not found in official list): {len(unmatched_user)}")

if unmatched_user:
    print("\n--- User conferences NOT in official 2026 list ---")
    for name, score in sorted(unmatched_user, key=lambda x: x[1], reverse=True):
        print(f"  [{score:.2f}] {name}")

# Find official items not in user list
official_matched = set()
for uc in user_confs:
    match, score = find_best_match(uc, official_conf_list, threshold=0.85)
    if match:
        official_matched.add(match[0])  # abbr

missing_from_user = [(abbr, full, rank, cat) for abbr, full, rank, cat in official_conf_list if abbr not in official_matched]
print(f"\nOfficial conferences NOT in user list: {len(missing_from_user)}")
if missing_from_user:
    print("\n--- Missing from user list ---")
    for abbr, full, rank, cat in missing_from_user:
        print(f"  [{rank}] {abbr}: {full} ({cat})")

# Count by rank
print("\n--- User conference count by estimated rank ---")
# Determine rank based on matched official entry
rank_counts = {"A": 0, "B": 0, "C": 0, "unknown": 0}
for uc in user_confs:
    match, score = find_best_match(uc, official_conf_list, threshold=0.85)
    if match:
        rank_counts[match[2]] += 1
    else:
        rank_counts["unknown"] += 1
print(f"  A: {rank_counts['A']} (official: 58)")
print(f"  B: {rank_counts['B']} (official: 132)")
print(f"  C: {rank_counts['C']} (official: 196)")
print(f"  unknown: {rank_counts['unknown']}")

# ============================================================
# JOURNAL ANALYSIS
# ============================================================
print("\n" + "=" * 80)
print("JOURNAL ANALYSIS")
print("=" * 80)

# Official 2026 journal list
official_jrn_list = [
    # 体系结构
    ("TOCS", "ACM Transactions on Computer Systems", "A", "体系结构"),
    ("TOS", "ACM Transactions on Storage", "A", "体系结构"),
    ("TCAD", "IEEE Transactions on Computer-Aided Design of Integrated Circuits and Systems", "A", "体系结构"),
    ("TC", "IEEE Transactions on Computers", "A", "体系结构"),
    ("TPDS", "IEEE Transactions on Parallel and Distributed Systems", "A", "体系结构"),
    ("TACO", "ACM Transactions on Architecture and Code Optimization", "A", "体系结构"),
    ("TAAS", "ACM Transactions on Autonomous and Adaptive Systems", "B", "体系结构"),
    ("TODAES", "ACM Transactions on Design Automation of Electronic Systems", "B", "体系结构"),
    ("TECS", "ACM Transactions on Embedded Computing Systems", "B", "体系结构"),
    ("TRETS", "ACM Transactions on Reconfigurable Technology and Systems", "B", "体系结构"),
    ("TVLSI", "IEEE Transactions on Very Large Scale Integration (VLSI) Systems", "B", "体系结构"),
    ("JPDC", "Journal of Parallel and Distributed Computing", "B", "体系结构"),
    ("JSA", "Journal of Systems Architecture: Embedded Software Design", "B", "体系结构"),
    ("Parallel Computing", "Parallel Computing", "B", "体系结构"),
    ("Performance Evaluation", "Performance Evaluation: An International Journal", "B", "体系结构"),
    ("TCC", "IEEE Transactions on Cloud Computing", "B", "体系结构"),
    ("JETC", "ACM Journal on Emerging Technologies in Computing Systems", "C", "体系结构"),
    ("CCPE", "Concurrency and Computation: Practice and Experience", "C", "体系结构"),
    ("DC", "Distributed Computing", "C", "体系结构"),
    ("FGCS", "Future Generation Computer Systems", "C", "体系结构"),
    ("Integration", "Integration, the VLSI Journal", "C", "体系结构"),
    ("JETTA", "Journal of Electronic Testing-Theory and Applications", "C", "体系结构"),
    ("JGC", "Journal of Grid Computing", "C", "体系结构"),
    ("RTS", "Real-Time Systems", "C", "体系结构"),
    ("TJSC", "The Journal of Supercomputing", "C", "体系结构"),
    ("TCASI", "IEEE Transactions on Circuits and Systems I: Regular Papers", "C", "体系结构"),
    ("CCF-THPC", "CCF Transactions on High Performance Computing", "C", "体系结构"),
    ("TSUSC", "IEEE Transactions on Sustainable Computing", "C", "体系结构"),
    # 计算机网络
    ("JSAC", "IEEE Journal on Selected Areas in Communications", "A", "计算机网络"),
    ("TMC", "IEEE Transactions on Mobile Computing", "A", "计算机网络"),
    ("TON", "IEEE/ACM Transactions on Networking", "A", "计算机网络"),
    ("TOIT", "ACM Transactions on Internet Technology", "B", "计算机网络"),
    ("TOMM", "ACM Transactions on Multimedia Computing, Communications and Applications", "B", "计算机网络"),
    ("TOSN", "ACM Transactions on Sensor Networks", "B", "计算机网络"),
    ("CN", "Computer Networks", "B", "计算机网络"),
    ("TCOM", "IEEE Transactions on Communications", "B", "计算机网络"),
    ("TWC", "IEEE Transactions on Wireless Communications", "B", "计算机网络"),
    ("Ad Hoc Networks", "Ad Hoc Networks", "C", "计算机网络"),
    ("CC", "Computer Communications", "C", "计算机网络"),
    ("TNSM", "IEEE Transactions on Network and Service Management", "C", "计算机网络"),
    ("IET Communications", "IET Communications", "C", "计算机网络"),
    ("JNCA", "Journal of Network and Computer Applications", "C", "计算机网络"),
    ("MONET", "Mobile Networks and Applications", "C", "计算机网络"),
    ("Networks", "Networks", "C", "计算机网络"),
    ("PPNA", "Peer-to-Peer Networking and Applications", "C", "计算机网络"),
    ("WCMC", "Wireless Communications and Mobile Computing", "C", "计算机网络"),
    ("Wireless Networks", "Wireless Networks", "C", "计算机网络"),
    ("IOT", "IEEE Internet of Things Journal", "C", "计算机网络"),
    ("TIOT", "ACM Transactions on Internet of Things", "C", "计算机网络"),
    # 信息安全
    ("TDSC", "IEEE Transactions on Dependable and Secure Computing", "A", "信息安全"),
    ("TIFS", "IEEE Transactions on Information Forensics and Security", "A", "信息安全"),
    ("Journal of Cryptology", "Journal of Cryptology", "A", "信息安全"),
    ("TOPS", "ACM Transactions on Privacy and Security", "B", "信息安全"),
    ("Computers & Security", "Computers & Security", "B", "信息安全"),
    ("Designs Codes and Cryptography", "Designs, Codes and Cryptography", "B", "信息安全"),
    ("JCS", "Journal of Computer Security", "B", "信息安全"),
    ("Cybersecurity", "Cybersecurity", "B", "信息安全"),
    ("CLSR", "Computer Law & Security Review", "C", "信息安全"),
    ("EURASIP JIS", "EURASIP Journal on Information Security", "C", "信息安全"),
    ("IET Information Security", "IET Information Security", "C", "信息安全"),
    ("IMCS", "Information and Computer Security", "C", "信息安全"),
    ("IJICS", "International Journal of Information and Computer Security", "C", "信息安全"),
    ("IJISP", "International Journal of Information Security and Privacy", "C", "信息安全"),
    ("JISA", "Journal of Information Security and Applications", "C", "信息安全"),
    ("SCN", "Security and Communication Networks", "C", "信息安全"),
    ("HCC", "High-Confidence Computing", "C", "信息安全"),
    # 软件工程
    ("TOPLAS", "ACM Transactions on Programming Languages and Systems", "A", "软件工程"),
    ("TOSEM", "ACM Transactions on Software Engineering and Methodology", "A", "软件工程"),
    ("TSE", "IEEE Transactions on Software Engineering", "A", "软件工程"),
    ("TSC", "IEEE Transactions on Services Computing", "A", "软件工程"),
    ("ASE", "Automated Software Engineering", "B", "软件工程"),
    ("ESE", "Empirical Software Engineering", "B", "软件工程"),
    ("IETS", "IET Software", "B", "软件工程"),
    ("IST", "Information and Software Technology", "B", "软件工程"),
    ("JFP", "Journal of Functional Programming", "B", "软件工程"),
    ("JSME", "Journal of Software: Evolution and Process", "B", "软件工程"),
    ("JSS", "Journal of Systems and Software", "B", "软件工程"),
    ("RE", "Requirements Engineering", "B", "软件工程"),
    ("SCP", "Science of Computer Programming", "B", "软件工程"),
    ("SoSyM", "Software and Systems Modeling", "B", "软件工程"),
    ("STVR", "Software Testing, Verification and Reliability", "B", "软件工程"),
    ("SPE", "Software: Practice and Experience", "B", "软件工程"),
    ("CL", "Computer Languages, Systems and Structures", "C", "软件工程"),
    ("IJSEKE", "International Journal of Software Engineering and Knowledge Engineering", "C", "软件工程"),
    ("STTT", "International Journal of Software Tools for Technology Transfer", "C", "软件工程"),
    ("JLAMP", "Journal of Logical and Algebraic Methods in Programming", "C", "软件工程"),
    ("JWE", "Journal of Web Engineering", "C", "软件工程"),
    ("SOCA", "Service Oriented Computing and Applications", "C", "软件工程"),
    ("SQJ", "Software Quality Journal", "C", "软件工程"),
    ("TPLP", "Theory and Practice of Logic Programming", "C", "软件工程"),
    ("PACMPL", "Proceedings of the ACM on Programming Languages", "C", "软件工程"),
    # 数据库
    ("TODS", "ACM Transactions on Database Systems", "A", "数据库"),
    ("TOIS", "ACM Transactions on Information Systems", "A", "数据库"),
    ("TKDE", "IEEE Transactions on Knowledge and Data Engineering", "A", "数据库"),
    ("VLDBJ", "The VLDB Journal", "A", "数据库"),
    ("TKDD", "ACM Transactions on Knowledge Discovery from Data", "B", "数据库"),
    ("TWEB", "ACM Transactions on the Web", "B", "数据库"),
    ("AEI", "Advanced Engineering Informatics", "B", "数据库"),
    ("DKE", "Data & Knowledge Engineering", "B", "数据库"),
    ("DMKD", "Data Mining and Knowledge Discovery", "B", "数据库"),
    ("EJIS", "European Journal of Information Systems", "B", "数据库"),
    ("GeoInformatica", "GeoInformatica", "B", "数据库"),
    ("IPM", "Information Processing and Management", "B", "数据库"),
    ("Information Sciences", "Information Sciences", "B", "数据库"),
    ("IS", "Information Systems", "B", "数据库"),
    ("JASIST", "Journal of the Association for Information Science and Technology", "B", "数据库"),
    ("JWS", "Journal of Web Semantics", "B", "数据库"),
    ("KAIS", "Knowledge and Information Systems", "B", "数据库"),
    ("DSE", "Data Science and Engineering", "B", "数据库"),
    ("DPD", "Distributed and Parallel Databases", "C", "数据库"),
    ("I&M", "Information & Management", "C", "数据库"),
    ("IPL", "Information Processing Letters", "C", "数据库"),
    ("IRJ", "Information Retrieval Journal", "C", "数据库"),
    ("IJCIS", "International Journal of Cooperative Information Systems", "C", "数据库"),
    ("IJGIS", "International Journal of Geographical Information Science", "C", "数据库"),
    ("IJIS", "International Journal of Intelligent Systems", "C", "数据库"),
    ("IJKM", "International Journal of Knowledge Management", "C", "数据库"),
    ("IJSWIS", "International Journal on Semantic Web and Information Systems", "C", "数据库"),
    ("JCIS", "Journal of Computer Information Systems", "C", "数据库"),
    ("JDM", "Journal of Database Management", "C", "数据库"),
    ("JGITM", "Journal of Global Information Technology Management", "C", "数据库"),
    ("JIIS", "Journal of Intelligent Information Systems", "C", "数据库"),
    ("JSIS", "The Journal of Strategic Information Systems", "C", "数据库"),
    ("TIST", "ACM Transactions on Intelligent Systems and Technology", "C", "数据库"),
    ("TORS", "ACM Transactions on Recommender Systems", "C", "数据库"),
    # 理论
    ("TIT", "IEEE Transactions on Information Theory", "A", "理论"),
    ("IANDC", "Information and Computation", "A", "理论"),
    ("SICOMP", "SIAM Journal on Computing", "A", "理论"),
    ("TALG", "ACM Transactions on Algorithms", "B", "理论"),
    ("TOCL", "ACM Transactions on Computational Logic", "B", "理论"),
    ("TOMS", "ACM Transactions on Mathematical Software", "B", "理论"),
    ("Algorithmica", "Algorithmica", "B", "理论"),
    ("CC", "Computational Complexity", "B", "理论"),
    ("FAC", "Formal Aspects of Computing", "B", "理论"),
    ("FMSD", "Formal Methods in System Design", "B", "理论"),
    ("INFORMS", "INFORMS Journal on Computing", "B", "理论"),
    ("JCSS", "Journal of Computer and System Sciences", "B", "理论"),
    ("JGO", "Journal of Global Optimization", "B", "理论"),
    ("JSC", "Journal of Symbolic Computation", "B", "理论"),
    ("MSCS", "Mathematical Structures in Computer Science", "B", "理论"),
    ("TCS", "Theoretical Computer Science", "B", "理论"),
    ("ACTA", "Acta Informatica", "C", "理论"),
    ("APAL", "Annals of Pure and Applied Logic", "C", "理论"),
    ("DAM", "Discrete Applied Mathematics", "C", "理论"),
    ("FUIN", "Fundamenta Informaticae", "C", "理论"),
    ("IPL", "Information Processing Letters", "C", "理论"),
    ("JCOMPLEXITY", "Journal of Complexity", "C", "理论"),
    ("LOGCOM", "Journal of Logic and Computation", "C", "理论"),
    ("JSL", "The Journal of Symbolic Logic", "C", "理论"),
    ("LMCS", "Logical Methods in Computer Science", "C", "理论"),
    ("SIDMA", "SIAM Journal on Discrete Mathematics", "C", "理论"),
    ("TOCS", "Theory of Computing Systems", "C", "理论"),
    ("TQC", "ACM Transactions in Quantum Computing", "C", "理论"),
    # 图形学多媒体
    ("TOG", "ACM Transactions on Graphics", "A", "图形学多媒体"),
    ("TIP", "IEEE Transactions on Image Processing", "A", "图形学多媒体"),
    ("TVCG", "IEEE Transactions on Visualization and Computer Graphics", "A", "图形学多媒体"),
    ("TMM", "IEEE Transactions on Multimedia", "A", "图形学多媒体"),
    ("TOMM", "ACM Transactions on Multimedia Computing, Communications and Applications", "B", "图形学多媒体"),
    ("CAGD", "Computer Aided Geometric Design", "B", "图形学多媒体"),
    ("CGF", "Computer Graphics Forum", "B", "图形学多媒体"),
    ("CAD", "Computer-Aided Design", "B", "图形学多媒体"),
    ("TCSVT", "IEEE Transactions on Circuits and Systems for Video Technology", "B", "图形学多媒体"),
    ("JASA", "The Journal of the Acoustical Society of America", "B", "图形学多媒体"),
    ("SIIMS", "SIAM Journal on Imaging Sciences", "B", "图形学多媒体"),
    ("SPECOM", "Speech Communication", "B", "图形学多媒体"),
    ("CVMJ", "Computational Visual Media", "B", "图形学多媒体"),
    ("CGTA", "Computational Geometry: Theory and Applications", "C", "图形学多媒体"),
    ("CAVW", "Computer Animation & Virtual Worlds", "C", "图形学多媒体"),
    ("C&G", "Computers & Graphics", "C", "图形学多媒体"),
    ("DCG", "Discrete & Computational Geometry", "C", "图形学多媒体"),
    ("SPL", "IEEE Signal Processing Letters", "C", "图形学多媒体"),
    ("IET-IPR", "IET Image Processing", "C", "图形学多媒体"),
    ("JVCIR", "Journal of Visual Communication and Image Representation", "C", "图形学多媒体"),
    ("MS", "Multimedia Systems", "C", "图形学多媒体"),
    ("MTA", "Multimedia Tools and Applications", "C", "图形学多媒体"),
    ("SIGPRO", "Signal Processing", "C", "图形学多媒体"),
    ("IMAGE", "Signal Processing: Image Communication", "C", "图形学多媒体"),
    ("TVC", "The Visual Computer", "C", "图形学多媒体"),
    ("VI", "Visual Informatics", "C", "图形学多媒体"),
    ("VRIH", "Virtual Reality & Intelligent Hardware", "C", "图形学多媒体"),
    ("GMOD", "Graphical Models", "C", "图形学多媒体"),
    # 人工智能
    ("AI", "Artificial Intelligence", "A", "人工智能"),
    ("TPAMI", "IEEE Transactions on Pattern Analysis and Machine Intelligence", "A", "人工智能"),
    ("IJCV", "International Journal of Computer Vision", "A", "人工智能"),
    ("JMLR", "Journal of Machine Learning Research", "A", "人工智能"),
    ("TAP", "ACM Transactions on Applied Perception", "B", "人工智能"),
    ("AAMAS", "Autonomous Agents and Multi-Agent Systems", "B", "人工智能"),
    ("Computational Linguistics", "Computational Linguistics", "B", "人工智能"),
    ("CVIU", "Computer Vision and Image Understanding", "B", "人工智能"),
    ("DKE", "Data & Knowledge Engineering", "B", "人工智能"),
    ("Evolutionary Computation", "Evolutionary Computation", "B", "人工智能"),
    ("TAC", "IEEE Transactions on Affective Computing", "B", "人工智能"),
    ("TASLP", "IEEE/ACM Transactions on Audio, Speech and Language Processing", "B", "人工智能"),
    ("TCYB", "IEEE Transactions on Cybernetics", "B", "人工智能"),
    ("TEC", "IEEE Transactions on Evolutionary Computation", "B", "人工智能"),
    ("TFS", "IEEE Transactions on Fuzzy Systems", "B", "人工智能"),
    ("TNNLS", "IEEE Transactions on Neural Networks and Learning Systems", "B", "人工智能"),
    ("IJAR", "International Journal of Approximate Reasoning", "B", "人工智能"),
    ("JAIR", "Journal of Artificial Intelligence Research", "B", "人工智能"),
    ("JAR", "Journal of Automated Reasoning", "B", "人工智能"),
    ("JSLHR", "Journal of Speech, Language, and Hearing Research", "B", "人工智能"),
    ("Machine Learning", "Machine Learning", "B", "人工智能"),
    ("Neural Computation", "Neural Computation", "B", "人工智能"),
    ("Neural Networks", "Neural Networks", "B", "人工智能"),
    ("PR", "Pattern Recognition", "B", "人工智能"),
    ("TACL", "Transactions of the Association for Computational Linguistics", "B", "人工智能"),
    ("TALLIP", "ACM Transactions on Asian and Low-Resource Language Information Processing", "C", "人工智能"),
    ("Applied Intelligence", "Applied Intelligence", "C", "人工智能"),
    ("AIM", "Artificial Intelligence in Medicine", "C", "人工智能"),
    ("Artificial Life", "Artificial Life", "C", "人工智能"),
    ("Computational Intelligence", "Computational Intelligence", "C", "人工智能"),
    ("Computer Speech & Language", "Computer Speech & Language", "C", "人工智能"),
    ("Connection Science", "Connection Science", "C", "人工智能"),
    ("DSS", "Decision Support Systems", "C", "人工智能"),
    ("EAAI", "Engineering Applications of Artificial Intelligence", "C", "人工智能"),
    ("Expert Systems", "Expert Systems with Applications", "C", "人工智能"),
    ("Fuzzy Sets and Systems", "Fuzzy Sets and Systems", "C", "人工智能"),
    ("TG", "IEEE Transactions on Games", "C", "人工智能"),
    ("IET-CVI", "IET Computer Vision", "C", "人工智能"),
    ("IET Signal Processing", "IET Signal Processing", "C", "人工智能"),
    ("IVC", "Image and Vision Computing", "C", "人工智能"),
    ("IDA", "Intelligent Data Analysis", "C", "人工智能"),
    ("IJCIA", "International Journal of Computational Intelligence and Applications", "C", "人工智能"),
    ("IJIS", "International Journal of Intelligent Systems", "C", "人工智能"),
    ("IJNS", "International Journal of Neural Systems", "C", "人工智能"),
    ("IJPRAI", "International Journal of Pattern Recognition and Artificial Intelligence", "C", "人工智能"),
    ("IJUFKS", "International Journal of Uncertainty, Fuzziness and Knowledge-Based Systems", "C", "人工智能"),
    ("IJDAR", "International Journal on Document Analysis and Recognition", "C", "人工智能"),
    ("JETAI", "Journal of Experimental and Theoretical Artificial Intelligence", "C", "人工智能"),
    ("KBS", "Knowledge-Based Systems", "C", "人工智能"),
    ("Machine Translation", "Machine Translation", "C", "人工智能"),
    ("Machine Vision and Applications", "Machine Vision and Applications", "C", "人工智能"),
    ("Natural Computing", "Natural Computing", "C", "人工智能"),
    ("NLE", "Natural Language Engineering", "C", "人工智能"),
    ("NCA", "Neural Computing and Applications", "C", "人工智能"),
    ("NPL", "Neural Processing Letters", "C", "人工智能"),
    ("Neurocomputing", "Neurocomputing", "C", "人工智能"),
    ("PAA", "Pattern Analysis and Applications", "C", "人工智能"),
    ("PRL", "Pattern Recognition Letters", "C", "人工智能"),
    ("Soft Computing", "Soft Computing", "C", "人工智能"),
    ("WI", "Web Intelligence", "C", "人工智能"),
    ("TIIS", "ACM Transactions on Interactive Intelligent Systems", "C", "人工智能"),
    ("TELO", "ACM Transactions on Evolutionary Learning and Optimization", "C", "人工智能"),
    ("JATS", "ACM Journal on Autonomous Transportation Systems", "C", "人工智能"),
    # 人机交互
    ("TOCHI", "ACM Transactions on Computer-Human Interaction", "A", "人机交互"),
    ("IJHCS", "International Journal of Human-Computer Studies", "A", "人机交互"),
    ("CSCW", "Computer Supported Cooperative Work", "B", "人机交互"),
    ("HCI", "Human-Computer Interaction", "B", "人机交互"),
    ("THMS", "IEEE Transactions on Human-Machine Systems", "B", "人机交互"),
    ("IWC", "Interacting with Computers", "B", "人机交互"),
    ("IJHCI", "International Journal of Human-Computer Interaction", "B", "人机交互"),
    ("UMUAI", "User Modeling and User-Adapted Interaction", "B", "人机交互"),
    ("TSMC", "IEEE Transactions on Systems, Man, and Cybernetics: Systems", "B", "人机交互"),
    ("CCF TPCI", "CCF Transactions on Pervasive Computing and Interaction", "B", "人机交互"),
    ("BIT", "Behaviour & Information Technology", "C", "人机交互"),
    ("PUC", "Personal and Ubiquitous Computing", "C", "人机交互"),
    ("PMC", "Pervasive and Mobile Computing", "C", "人机交互"),
    ("PACMHCI", "Proceedings of the ACM on Human-Computer Interaction", "C", "人机交互"),
    ("THRI", "ACM Transactions on Human-Robot Interaction", "C", "人机交互"),
    # 交叉综合
    ("JACM", "Journal of the ACM", "A", "交叉综合"),
    ("Proc. IEEE", "Proceedings of the IEEE", "A", "交叉综合"),
    ("SCIS", "Science China Information Sciences", "A", "交叉综合"),
    ("Bioinformatics", "Bioinformatics", "A", "交叉综合"),
    ("Briefings in Bioinformatics", "Briefings in Bioinformatics", "B", "交叉综合"),
    ("Cognition", "Cognition", "B", "交叉综合"),
    ("TASE", "IEEE Transactions on Automation Science and Engineering", "B", "交叉综合"),
    ("TGARS", "IEEE Transactions on Geoscience and Remote Sensing", "B", "交叉综合"),
    ("TITS", "IEEE Transactions on Intelligent Transportation Systems", "B", "交叉综合"),
    ("TMI", "IEEE Transactions on Medical Imaging", "B", "交叉综合"),
    ("TR", "IEEE Transactions on Robotics", "B", "交叉综合"),
    ("TCBB", "IEEE/ACM Transactions on Computational Biology and Bioinformatics", "B", "交叉综合"),
    ("JCST", "Journal of Computer Science and Technology", "B", "交叉综合"),
    ("JAMIA", "Journal of the American Medical Informatics Association", "B", "交叉综合"),
    ("PLOS Computational Biology", "PLOS Computational Biology", "B", "交叉综合"),
    ("The Computer Journal", "The Computer Journal", "B", "交叉综合"),
    ("WWW Journal", "World Wide Web", "B", "交叉综合"),
    ("FCS", "Frontiers of Computer Science", "B", "交叉综合"),
    ("BCRA", "Blockchain: Research and Applications", "B", "交叉综合"),
    ("BMC Bioinformatics", "BMC Bioinformatics", "C", "交叉综合"),
    ("Cybernetics and Systems", "Cybernetics and Systems", "C", "交叉综合"),
    ("GRSL", "IEEE Geoscience and Remote Sensing Letters", "C", "交叉综合"),
    ("JBHI", "IEEE Journal of Biomedical and Health Informatics", "C", "交叉综合"),
    ("TBD", "IEEE Transactions on Big Data", "C", "交叉综合"),
    ("IET Intelligent Transport Systems", "IET Intelligent Transport Systems", "C", "交叉综合"),
    ("JBI", "Journal of Biomedical Informatics", "C", "交叉综合"),
    ("Medical Image Analysis", "Medical Image Analysis", "C", "交叉综合"),
    ("TII", "IEEE Transactions on Industrial Informatics", "C", "交叉综合"),
    ("TCPS", "ACM Transactions on Cyber-Physical Systems", "C", "交叉综合"),
    ("TOCE", "ACM Transactions on Computing Education", "C", "交叉综合"),
    ("FITEE", "Frontiers of Information Technology & Electronic Engineering", "C", "交叉综合"),
    ("TCSS", "IEEE Transactions on Computational Social Systems", "C", "交叉综合"),
    ("TRel", "IEEE Transactions on Reliability", "C", "交叉综合"),
    ("HEALTH", "ACM Transactions on Computing for Healthcare", "C", "交叉综合"),
    ("ACM DLT", "ACM Distributed Ledger Technologies: Research and Practice", "C", "交叉综合"),
    # Extra journals in 2026
    ("IACR TCHES", "IACR Transactions on Cryptographic Hardware and Embedded Systems", "C", "交叉综合"),
    ("IACR ToSC", "IACR Transactions on Symmetric Cryptology", "C", "交叉综合"),
]

print(f"\nOfficial journals: {len(official_jrn_list)}")
print(f"User journals: {len(user_jrns)}")

# Match
matched_j = []
unmatched_user_j = []
for uj in user_jrns:
    match, score = find_best_match(uj, official_jrn_list, threshold=0.85)
    if match:
        matched_j.append((uj, match, score))
    else:
        unmatched_user_j.append((uj, score))

print(f"\nMatched journals: {len(matched_j)}")
print(f"Unmatched (user items not found in official list): {len(unmatched_user_j)}")

if unmatched_user_j:
    print("\n--- User journals NOT in official 2026 list ---")
    for name, score in sorted(unmatched_user_j, key=lambda x: x[1], reverse=True):
        print(f"  [{score:.2f}] {name}")

# Find official items not in user list
official_jrn_matched = set()
for uj in user_jrns:
    match, score = find_best_match(uj, official_jrn_list, threshold=0.85)
    if match:
        official_jrn_matched.add(match[0])

missing_j = [(abbr, full, rank, cat) for abbr, full, rank, cat in official_jrn_list if abbr not in official_jrn_matched]
print(f"\nOfficial journals NOT in user list: {len(missing_j)}")
if missing_j:
    print("\n--- Missing journals ---")
    for abbr, full, rank, cat in missing_j:
        print(f"  [{rank}] {abbr}: {full} ({cat})")

# Duplicate analysis
print("\n=== JOURNAL DUPLICATES IN USER DATA ===")
from collections import Counter
jrn_name_counts = Counter(user_jrns)
dupes = {name: count for name, count in jrn_name_counts.items() if count > 1}
if dupes:
    for name, count in dupes.items():
        print(f"  '{name}' appears {count} times")
        # Check if this is legit (appears in multiple categories)
        official_matches = [(abbr, full, rank, cat) for abbr, full, rank, cat in official_jrn_list if normalize(full) == normalize(name)]
        if len(official_matches) > 1:
            print(f"    -> Legitimate: appears in categories: {[m[3] for m in official_matches]}")
        else:
            print(f"    -> LIKELY DUPLICATE ERROR")
else:
    print("  No duplicate journal names found")

print("\nDone!")
