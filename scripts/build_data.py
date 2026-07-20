#!/usr/bin/env python3
"""
Build clean JSON data files from Excel source.
Uses full official names for precise matching (verified 385/387 accuracy).
"""
import json
import re
import sys
from pathlib import Path
from collections import OrderedDict
from difflib import SequenceMatcher

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)

# ============================================================
# OFFICIAL 2026 CCF FULL NAMES (from detailed_comparison.py verified matching)
# ============================================================
# Format: (abbreviation, full_official_name, rank, category_zh, category_en)

CCF_CONF_FULL = [
    # === 体系结构 A (11) ===
    ("PPoPP", "ACM SIGPLAN Symposium on Principles & Practice of Parallel Programming", "A", "计算机体系结构", "Computer Architecture"),
    ("FAST", "USENIX Conference on File and Storage Technologies", "A", "计算机体系结构", "Computer Architecture"),
    ("DAC", "Design Automation Conference", "A", "计算机体系结构", "Computer Architecture"),
    ("HPCA", "IEEE International Symposium on High Performance Computer Architecture", "A", "计算机体系结构", "Computer Architecture"),
    ("MICRO", "IEEE/ACM International Symposium on Microarchitecture", "A", "计算机体系结构", "Computer Architecture"),
    ("SC", "International Conference for High Performance Computing, Networking, Storage, and Analysis", "A", "计算机体系结构", "Computer Architecture"),
    ("ASPLOS", "International Conference on Architectural Support for Programming Languages and Operating Systems", "A", "计算机体系结构", "Computer Architecture"),
    ("ISCA", "International Symposium on Computer Architecture", "A", "计算机体系结构", "Computer Architecture"),
    ("ATC", "USENIX Annual Technical Conference", "A", "计算机体系结构", "Computer Architecture"),
    ("EuroSys", "European Conference on Computer Systems", "A", "计算机体系结构", "Computer Architecture"),
    ("HPDC", "International ACM Symposium on High-Performance Parallel and Distributed Computing", "A", "计算机体系结构", "Computer Architecture"),
    # === 体系结构 B (26) ===
    ("SoCC", "ACM Symposium on Cloud Computing", "B", "计算机体系结构", "Computer Architecture"),
    ("SPAA", "ACM Symposium on Parallelism in Algorithms and Architectures", "B", "计算机体系结构", "Computer Architecture"),
    ("PODC", "ACM Symposium on Principles of Distributed Computing", "B", "计算机体系结构", "Computer Architecture"),
    ("FPGA", "ACM/SIGDA International Symposium on Field-Programmable Gate Arrays", "B", "计算机体系结构", "Computer Architecture"),
    ("CGO", "International Symposium on Code Generation and Optimization", "B", "计算机体系结构", "Computer Architecture"),
    ("DATE", "Design, Automation & Test in Europe", "B", "计算机体系结构", "Computer Architecture"),
    ("HOT CHIPS", "Hot Chips: A Symposium on High Performance Chips", "B", "计算机体系结构", "Computer Architecture"),
    ("CLUSTER", "IEEE International Conference on Cluster Computing", "B", "计算机体系结构", "Computer Architecture"),
    ("ICCD", "International Conference on Computer Design", "B", "计算机体系结构", "Computer Architecture"),
    ("ICCAD", "International Conference on Computer-Aided Design", "B", "计算机体系结构", "Computer Architecture"),
    ("ICDCS", "IEEE International Conference on Distributed Computing Systems", "B", "计算机体系结构", "Computer Architecture"),
    ("CODES+ISSS", "International Conference on Hardware/Software Co-design and System Synthesis", "B", "计算机体系结构", "Computer Architecture"),
    ("HiPEAC", "International Conference on High Performance and Embedded Architectures and Compilers", "B", "计算机体系结构", "Computer Architecture"),
    ("SIGMETRICS", "International Conference on Measurement and Modeling of Computer Systems", "B", "计算机体系结构", "Computer Architecture"),
    ("PACT", "International Conference on Parallel Architectures and Compilation Techniques", "B", "计算机体系结构", "Computer Architecture"),
    ("ICPP", "International Conference on Parallel Processing", "B", "计算机体系结构", "Computer Architecture"),
    ("ICS", "International Conference on Supercomputing", "B", "计算机体系结构", "Computer Architecture"),
    ("VEE", "International Conference on Virtual Execution Environments", "B", "计算机体系结构", "Computer Architecture"),
    ("IPDPS", "IEEE International Parallel & Distributed Processing Symposium", "B", "计算机体系结构", "Computer Architecture"),
    ("Performance", "International Symposium on Computer Performance, Modeling, Measurements and Evaluation", "B", "计算机体系结构", "Computer Architecture"),
    ("ITC", "International Test Conference", "B", "计算机体系结构", "Computer Architecture"),
    ("LISA", "Large Installation System Administration Conference", "B", "计算机体系结构", "Computer Architecture"),
    ("MSST", "Mass Storage Systems and Technologies", "B", "计算机体系结构", "Computer Architecture"),
    ("RTAS", "IEEE Real-Time and Embedded Technology and Applications Symposium", "B", "计算机体系结构", "Computer Architecture"),
    ("Euro-Par", "European Conference on Parallel and Distributed Computing", "B", "计算机体系结构", "Computer Architecture"),
    ("ISCAS", "IEEE International Symposium on Circuits and Systems", "B", "计算机体系结构", "Computer Architecture"),
    # === 体系结构 C (30) ===
    ("CF", "ACM International Conference on Computing Frontiers", "C", "计算机体系结构", "Computer Architecture"),
    ("SYSTOR", "ACM International Systems and Storage Conference", "C", "计算机体系结构", "Computer Architecture"),
    ("NOCS", "ACM/IEEE International Symposium on Networks-on-Chip", "C", "计算机体系结构", "Computer Architecture"),
    ("ASAP", "IEEE International Conference on Application-Specific Systems, Architectures, and Processors", "C", "计算机体系结构", "Computer Architecture"),
    ("ASP-DAC", "Asia and South Pacific Design Automation Conference", "C", "计算机体系结构", "Computer Architecture"),
    ("ETS", "IEEE European Test Symposium", "C", "计算机体系结构", "Computer Architecture"),
    ("FPL", "International Conference on Field-Programmable Logic and Applications", "C", "计算机体系结构", "Computer Architecture"),
    ("FCCM", "IEEE Symposium on Field-Programmable Custom Computing Machines", "C", "计算机体系结构", "Computer Architecture"),
    ("GLSVLSI", "Great Lakes Symposium on VLSI", "C", "计算机体系结构", "Computer Architecture"),
    ("ATS", "IEEE Asian Test Symposium", "C", "计算机体系结构", "Computer Architecture"),
    ("HPCC", "IEEE International Conference on High Performance Computing and Communications", "C", "计算机体系结构", "Computer Architecture"),
    ("HiPC", "IEEE International Conference on High Performance Computing, Data and Analytics", "C", "计算机体系结构", "Computer Architecture"),
    ("MASCOTS", "International Symposium on Modeling, Analysis, and Simulation of Computer and Telecommunication Systems", "C", "计算机体系结构", "Computer Architecture"),
    ("ISPA", "IEEE International Symposium on Parallel and Distributed Processing with Applications", "C", "计算机体系结构", "Computer Architecture"),
    ("CCGRID", "IEEE/ACM International Symposium on Cluster, Cloud and Grid Computing", "C", "计算机体系结构", "Computer Architecture"),
    ("NPC", "IFIP International Conference on Network and Parallel Computing", "C", "计算机体系结构", "Computer Architecture"),
    ("ICA3PP", "International Conference on Algorithms and Architectures for Parallel Processing", "C", "计算机体系结构", "Computer Architecture"),
    ("CASES", "International Conference on Compilers, Architectures, and Synthesis for Embedded Systems", "C", "计算机体系结构", "Computer Architecture"),
    ("FPT", "International Conference on Field-Programmable Technology", "C", "计算机体系结构", "Computer Architecture"),
    ("ICPADS", "International Conference on Parallel and Distributed Systems", "C", "计算机体系结构", "Computer Architecture"),
    ("ISLPED", "International Symposium on Low Power Electronics and Design", "C", "计算机体系结构", "Computer Architecture"),
    ("ISPD", "International Symposium on Physical Design", "C", "计算机体系结构", "Computer Architecture"),
    ("HOTI", "IEEE Symposium on High-Performance Interconnects", "C", "计算机体系结构", "Computer Architecture"),
    ("VTS", "IEEE VLSI Test Symposium", "C", "计算机体系结构", "Computer Architecture"),
    ("ITC-Asia", "International Test Conference in Asia", "C", "计算机体系结构", "Computer Architecture"),
    ("SEC", "ACM/IEEE Symposium on Edge Computing", "C", "计算机体系结构", "Computer Architecture"),
    ("NAS", "International Conference on Networking, Architecture and Storages", "C", "计算机体系结构", "Computer Architecture"),
    ("HotStorage", "HotStorage", "C", "计算机体系结构", "Computer Architecture"),
    ("APPT", "International Symposium on Advanced Parallel Processing Technology", "C", "计算机体系结构", "Computer Architecture"),
    ("JCC", "International Conference on JointCloud Computing", "C", "计算机体系结构", "Computer Architecture"),

    # === 计算机网络 A (4) ===
    ("SIGCOMM", "ACM International Conference on Applications, Technologies, Architectures, and Protocols for Computer Communication", "A", "计算机网络", "Computer Networks"),
    ("MobiCom", "ACM International Conference on Mobile Computing and Networking", "A", "计算机网络", "Computer Networks"),
    ("INFOCOM", "IEEE International Conference on Computer Communications", "A", "计算机网络", "Computer Networks"),
    ("NSDI", "Symposium on Network System Design and Implementation", "A", "计算机网络", "Computer Networks"),
    # === 计算机网络 B (10) ===
    ("SenSys", "ACM Conference on Embedded Networked Sensor Systems", "B", "计算机网络", "Computer Networks"),
    ("CoNEXT", "ACM International Conference on Emerging Networking Experiments and Technologies", "B", "计算机网络", "Computer Networks"),
    ("SECON", "IEEE International Conference on Sensing, Communication, and Networking", "B", "计算机网络", "Computer Networks"),
    ("IPSN", "International Conference on Information Processing in Sensor Networks", "B", "计算机网络", "Computer Networks"),
    ("MobiSys", "ACM International Conference on Mobile Systems, Applications, and Services", "B", "计算机网络", "Computer Networks"),
    ("ICNP", "IEEE International Conference on Network Protocols", "B", "计算机网络", "Computer Networks"),
    ("MobiHoc", "International Symposium on Theory, Algorithmic Foundations, and Protocol Design for Mobile Networks and Mobile Computing", "B", "计算机网络", "Computer Networks"),
    ("NOSSDAV", "International Workshop on Network and Operating System Support for Digital Audio and Video", "B", "计算机网络", "Computer Networks"),
    ("IWQoS", "IEEE/ACM International Workshop on Quality of Service", "B", "计算机网络", "Computer Networks"),
    ("IMC", "ACM Internet Measurement Conference", "B", "计算机网络", "Computer Networks"),
    # === 计算机网络 C (20) ===
    ("ANCS", "ACM/IEEE Symposium on Architectures for Networking and Communication Systems", "C", "计算机网络", "Computer Networks"),
    ("APNOMS", "Asia-Pacific Network Operations and Management Symposium", "C", "计算机网络", "Computer Networks"),
    ("FORTE", "International Conference on Formal Techniques for Distributed Objects, Components, and Systems", "C", "计算机网络", "Computer Networks"),
    ("LCN", "IEEE Conference on Local Computer Networks", "C", "计算机网络", "Computer Networks"),
    ("GLOBECOM", "IEEE Global Communications Conference", "C", "计算机网络", "Computer Networks"),
    ("ICC", "IEEE International Conference on Communications", "C", "计算机网络", "Computer Networks"),
    ("ICCCN", "IEEE International Conference on Computer Communications and Networks", "C", "计算机网络", "Computer Networks"),
    ("MASS", "IEEE International Conference on Mobile Adhoc and Sensor Systems", "C", "计算机网络", "Computer Networks"),
    ("P2P", "IEEE International Conference on Peer-to-Peer Computing", "C", "计算机网络", "Computer Networks"),
    ("IPCCC", "IEEE International Performance Computing and Communications Conference", "C", "计算机网络", "Computer Networks"),
    ("WoWMoM", "IEEE International Symposium on a World of Wireless, Mobile and Multimedia Networks", "C", "计算机网络", "Computer Networks"),
    ("ISCC", "IEEE Symposium on Computers and Communications", "C", "计算机网络", "Computer Networks"),
    ("WCNC", "IEEE Wireless Communications and Networking Conference", "C", "计算机网络", "Computer Networks"),
    ("Networking", "IFIP International Conferences on Networking", "C", "计算机网络", "Computer Networks"),
    ("IM", "IFIP/IEEE International Symposium on Integrated Network Management", "C", "计算机网络", "Computer Networks"),
    ("MSN", "International Conference on Mobility, Sensing and Networking", "C", "计算机网络", "Computer Networks"),
    ("MSWiM", "International Conference on Modeling, Analysis and Simulation of Wireless and Mobile Systems", "C", "计算机网络", "Computer Networks"),
    ("WASA", "International Conference on Wireless Artificial Intelligent Computing Systems and Applications", "C", "计算机网络", "Computer Networks"),
    ("HotNets", "ACM Workshop on Hot Topics in Networks", "C", "计算机网络", "Computer Networks"),
    ("APNet", "Asia-Pacific Workshop on Networking", "C", "计算机网络", "Computer Networks"),

    # === 网络与信息安全 A (6) ===
    ("CCS", "ACM Conference on Computer and Communications Security", "A", "网络与信息安全", "Network & Information Security"),
    ("EUROCRYPT", "International Conference on the Theory and Applications of Cryptographic Techniques", "A", "网络与信息安全", "Network & Information Security"),
    ("S&P", "IEEE Symposium on Security and Privacy", "A", "网络与信息安全", "Network & Information Security"),
    ("CRYPTO", "International Cryptology Conference", "A", "网络与信息安全", "Network & Information Security"),
    ("USENIX Security", "USENIX Security Symposium", "A", "网络与信息安全", "Network & Information Security"),
    ("NDSS", "Network and Distributed System Security Symposium", "A", "网络与信息安全", "Network & Information Security"),
    # === 信息安全 B (11) ===
    ("ACSAC", "Annual Computer Security Applications Conference", "B", "网络与信息安全", "Network & Information Security"),
    ("ASIACRYPT", "Annual International Conference on the Theory and Application of Cryptology and Information Security", "B", "网络与信息安全", "Network & Information Security"),
    ("ESORICS", "European Symposium on Research in Computer Security", "B", "网络与信息安全", "Network & Information Security"),
    ("FSE", "Fast Software Encryption", "B", "网络与信息安全", "Network & Information Security"),
    ("CSFW", "IEEE Computer Security Foundations Workshop", "B", "网络与信息安全", "Network & Information Security"),
    ("SRDS", "IEEE International Symposium on Reliable Distributed Systems", "B", "网络与信息安全", "Network & Information Security"),
    ("CHES", "International Conference on Cryptographic Hardware and Embedded Systems", "B", "网络与信息安全", "Network & Information Security"),
    ("DSN", "International Conference on Dependable Systems and Networks", "B", "网络与信息安全", "Network & Information Security"),
    ("RAID", "International Symposium on Recent Advances in Intrusion Detection", "B", "网络与信息安全", "Network & Information Security"),
    ("PKC", "International Workshop on Practice and Theory in Public Key Cryptography", "B", "网络与信息安全", "Network & Information Security"),
    ("TCC", "Theory of Cryptography Conference", "B", "网络与信息安全", "Network & Information Security"),
    # === 信息安全 C (29) ===
    ("WiSec", "ACM Conference on Security and Privacy in Wireless and Mobile Networks", "C", "网络与信息安全", "Network & Information Security"),
    ("SACMAT", "ACM Symposium on Access Control Models and Technologies", "C", "网络与信息安全", "Network & Information Security"),
    ("DRM", "ACM Workshop on Digital Rights Management", "C", "网络与信息安全", "Network & Information Security"),
    ("IH&MMSec", "ACM Workshop on Information Hiding and Multimedia Security", "C", "网络与信息安全", "Network & Information Security"),
    ("ACNS", "International Conference on Applied Cryptography and Network Security", "C", "网络与信息安全", "Network & Information Security"),
    ("AsiaCCS", "ACM Asia Conference on Computer and Communications Security", "C", "网络与信息安全", "Network & Information Security"),
    ("ACISP", "Australasia Conference on Information Security and Privacy", "C", "网络与信息安全", "Network & Information Security"),
    ("CT-RSA", "The Cryptographer's Track at RSA Conference", "C", "网络与信息安全", "Network & Information Security"),
    ("DIMVA", "Conference on Detection of Intrusions and Malware & Vulnerability Assessment", "C", "网络与信息安全", "Network & Information Security"),
    ("DFRWS", "Digital Forensic Research Workshop", "C", "网络与信息安全", "Network & Information Security"),
    ("FC", "Financial Cryptography and Data Security", "C", "网络与信息安全", "Network & Information Security"),
    ("TrustCom", "IEEE International Conference on Trust, Security and Privacy in Computing and Communications", "C", "网络与信息安全", "Network & Information Security"),
    ("IFIP SEC", "IFIP International Information Security Conference", "C", "网络与信息安全", "Network & Information Security"),
    ("IFIP WG 11.9", "IFIP Working Group 11.9 International Conference on Digital Forensics", "C", "网络与信息安全", "Network & Information Security"),
    ("ISC", "Information Security Conference", "C", "网络与信息安全", "Network & Information Security"),
    ("ICDF2C", "International Conference on Digital Forensics & Cyber Crime", "C", "网络与信息安全", "Network & Information Security"),
    ("ICICS", "International Conference on Information and Communications Security", "C", "网络与信息安全", "Network & Information Security"),
    ("SecureComm", "International Conference on Security and Privacy in Communication Networks", "C", "网络与信息安全", "Network & Information Security"),
    ("NSPW", "New Security Paradigms Workshop", "C", "网络与信息安全", "Network & Information Security"),
    ("PAM", "Passive and Active Measurement Conference", "C", "网络与信息安全", "Network & Information Security"),
    ("PETS", "Privacy Enhancing Technologies Symposium", "C", "网络与信息安全", "Network & Information Security"),
    ("SAC", "Selected Areas in Cryptography", "C", "网络与信息安全", "Network & Information Security"),
    ("SOUPS", "Symposium On Usable Privacy and Security", "C", "网络与信息安全", "Network & Information Security"),
    ("HotSec", "USENIX Workshop on Hot Topics in Security", "C", "网络与信息安全", "Network & Information Security"),
    ("EuroS&P", "IEEE European Symposium on Security and Privacy", "C", "网络与信息安全", "Network & Information Security"),
    ("Inscrypt", "International Conference on Information Security and Cryptology", "C", "网络与信息安全", "Network & Information Security"),
    ("CODASPY", "Conference on Data and Application Security and Privacy", "C", "网络与信息安全", "Network & Information Security"),
    ("BlockSys", "International Conference on Blockchain, Artificial Intelligence, and Trustworthy Systems", "C", "网络与信息安全", "Network & Information Security"),
    ("CSCloud", "International Conference on Cyber Security and Cloud Computing", "C", "网络与信息安全", "Network & Information Security"),

    # === 软件工程 A (10) ===
    ("PLDI", "ACM SIGPLAN Conference on Programming Language Design and Implementation", "A", "软件工程", "Software Engineering"),
    ("POPL", "ACM SIGPLAN-SIGACT Symposium on Principles of Programming Languages", "A", "软件工程", "Software Engineering"),
    ("FSE", "ACM International Conference on the Foundations of Software Engineering", "A", "软件工程", "Software Engineering"),
    ("SOSP", "ACM Symposium on Operating Systems Principles", "A", "软件工程", "Software Engineering"),
    ("OOPSLA", "Conference on Object-Oriented Programming Systems, Languages, and Applications", "A", "软件工程", "Software Engineering"),
    ("ASE", "International Conference on Automated Software Engineering", "A", "软件工程", "Software Engineering"),
    ("ICSE", "International Conference on Software Engineering", "A", "软件工程", "Software Engineering"),
    ("ISSTA", "International Symposium on Software Testing and Analysis", "A", "软件工程", "Software Engineering"),
    ("OSDI", "USENIX Symposium on Operating Systems Design and Implementation", "A", "软件工程", "Software Engineering"),
    ("FM", "International Symposium on Formal Methods", "A", "软件工程", "Software Engineering"),
    # === 软件工程 B (20) ===
    ("ECOOP", "European Conference on Object-Oriented Programming", "B", "软件工程", "Software Engineering"),
    ("ETAPS", "European Joint Conferences on Theory and Practice of Software", "B", "软件工程", "Software Engineering"),
    ("ICPC", "IEEE International Conference on Program Comprehension", "B", "软件工程", "Software Engineering"),
    ("RE", "IEEE International Requirements Engineering Conference", "B", "软件工程", "Software Engineering"),
    ("CAiSE", "International Conference on Advanced Information Systems Engineering", "B", "软件工程", "Software Engineering"),
    ("ICFP", "ACM SIGPLAN International Conference on Function Programming", "B", "软件工程", "Software Engineering"),
    ("LCTES", "ACM SIGPLAN/SIGBED International Conference on Languages, Compilers and Tools for Embedded Systems", "B", "软件工程", "Software Engineering"),
    ("MoDELS", "ACM/IEEE International Conference on Model Driven Engineering Languages and Systems", "B", "软件工程", "Software Engineering"),
    ("CP", "International Conference on Principles and Practice of Constraint Programming", "B", "软件工程", "Software Engineering"),
    ("ICSOC", "International Conference on Service Oriented Computing", "B", "软件工程", "Software Engineering"),
    ("SANER", "IEEE International Conference on Software Analysis, Evolution, and Reengineering", "B", "软件工程", "Software Engineering"),
    ("ICSME", "International Conference on Software Maintenance and Evolution", "B", "软件工程", "Software Engineering"),
    ("VMCAI", "International Conference on Verification, Model Checking and Abstract Interpretation", "B", "软件工程", "Software Engineering"),
    ("ICWS", "IEEE International Conference on Web Services", "B", "软件工程", "Software Engineering"),
    ("Middleware", "International Middleware Conference", "B", "软件工程", "Software Engineering"),
    ("SAS", "International Static Analysis Symposium", "B", "软件工程", "Software Engineering"),
    ("ESEM", "International Symposium on Empirical Software Engineering and Measurement", "B", "软件工程", "Software Engineering"),
    ("ISSRE", "IEEE International Symposium on Software Reliability Engineering", "B", "软件工程", "Software Engineering"),
    ("HotOS", "USENIX Workshop on Hot Topics in Operating Systems", "B", "软件工程", "Software Engineering"),
    ("CC", "International Conference on Compiler Construction", "B", "软件工程", "Software Engineering"),
    # === 软件工程 C (27) ===
    ("PEPM", "ACM SIGPLAN Workshop on Partial Evaluation and Program Manipulation", "C", "软件工程", "Software Engineering"),
    ("PASTE", "ACM SIGPLAN-SIGSOFT Workshop on Program Analysis for Software Tools and Engineering", "C", "软件工程", "Software Engineering"),
    ("APLAS", "Asian Symposium on Programming Languages and Systems", "C", "软件工程", "Software Engineering"),
    ("APSEC", "Asia-Pacific Software Engineering Conference", "C", "软件工程", "Software Engineering"),
    ("EASE", "International Conference on Evaluation and Assessment in Software Engineering", "C", "软件工程", "Software Engineering"),
    ("ICECCS", "International Conference on Engineering of Complex Computer Systems", "C", "软件工程", "Software Engineering"),
    ("ICST", "IEEE International Conference on Software Testing, Verification and Validation", "C", "软件工程", "Software Engineering"),
    ("ISPASS", "IEEE International Symposium on Performance Analysis of Systems and Software", "C", "软件工程", "Software Engineering"),
    ("SCAM", "IEEE International Working Conference on Source Code Analysis and Manipulation", "C", "软件工程", "Software Engineering"),
    ("COMPSAC", "International Computer Software and Applications Conference", "C", "软件工程", "Software Engineering"),
    ("ICFEM", "International Conference on Formal Engineering Methods", "C", "软件工程", "Software Engineering"),
    ("SSE", "IEEE International Conference on Software Services Engineering", "C", "软件工程", "Software Engineering"),
    ("ICSSP", "International Conference on Software and System Process", "C", "软件工程", "Software Engineering"),
    ("SEKE", "International Conference on Software Engineering and Knowledge Engineering", "C", "软件工程", "Software Engineering"),
    ("QRS", "International Conference on Software Quality, Reliability and Security", "C", "软件工程", "Software Engineering"),
    ("ICSR", "International Conference on Software Reuse", "C", "软件工程", "Software Engineering"),
    ("ICWE", "International Conference on Web Engineering", "C", "软件工程", "Software Engineering"),
    ("SPIN", "International Symposium on Model Checking of Software", "C", "软件工程", "Software Engineering"),
    ("ATVA", "International Symposium on Automated Technology for Verification and Analysis", "C", "软件工程", "Software Engineering"),
    ("LOPSTR", "International Symposium on Logic-based Program Synthesis and Transformation", "C", "软件工程", "Software Engineering"),
    ("TASE", "Theoretical Aspects of Software Engineering Conference", "C", "软件工程", "Software Engineering"),
    ("MSR", "Mining Software Repositories", "C", "软件工程", "Software Engineering"),
    ("REFSQ", "Requirements Engineering: Foundation for Software Quality", "C", "软件工程", "Software Engineering"),
    ("WICSA", "Working IEEE/IFIP Conference on Software Architecture", "C", "软件工程", "Software Engineering"),
    ("Internetware", "Asia-Pacific Symposium on Internetware", "C", "软件工程", "Software Engineering"),
    ("RV", "International Conference on Runtime Verification", "C", "软件工程", "Software Engineering"),
    ("MEMOCODE", "International Conference on Formal Methods and Models for Co-Design", "C", "软件工程", "Software Engineering"),

    # === 数据库 A (5) ===
    ("SIGMOD", "ACM SIGMOD Conference", "A", "数据库/数据挖掘", "Databases & Data Mining"),
    ("SIGKDD", "ACM SIGKDD Conference on Knowledge Discovery and Data Mining", "A", "数据库/数据挖掘", "Databases & Data Mining"),
    ("ICDE", "IEEE International Conference on Data Engineering", "A", "数据库/数据挖掘", "Databases & Data Mining"),
    ("SIGIR", "International ACM SIGIR Conference on Research and Development in Information Retrieval", "A", "数据库/数据挖掘", "Databases & Data Mining"),
    ("VLDB", "International Conference on Very Large Data Bases", "A", "数据库/数据挖掘", "Databases & Data Mining"),
    # === 数据库 B (13) ===
    ("CIKM", "ACM International Conference on Information and Knowledge Management", "B", "数据库/数据挖掘", "Databases & Data Mining"),
    ("WSDM", "ACM International Conference on Web Search and Data Mining", "B", "数据库/数据挖掘", "Databases & Data Mining"),
    ("PODS", "ACM SIGMOD-SIGACT-SIGAI Symposium on Principles of Database Systems", "B", "数据库/数据挖掘", "Databases & Data Mining"),
    ("DASFAA", "International Conference on Database Systems for Advanced Applications", "B", "数据库/数据挖掘", "Databases & Data Mining"),
    ("ECML-PKDD", "European Conference on Machine Learning and Principles and Practice of Knowledge Discovery in Databases", "B", "数据库/数据挖掘", "Databases & Data Mining"),
    ("ISWC", "IEEE International Semantic Web Conference", "B", "数据库/数据挖掘", "Databases & Data Mining"),
    ("ICDM", "IEEE International Conference on Data Mining", "B", "数据库/数据挖掘", "Databases & Data Mining"),
    ("ICDT", "International Conference on Database Theory", "B", "数据库/数据挖掘", "Databases & Data Mining"),
    ("EDBT", "International Conference on Extending Database Technology", "B", "数据库/数据挖掘", "Databases & Data Mining"),
    ("CIDR", "Conference on Innovative Data Systems Research", "B", "数据库/数据挖掘", "Databases & Data Mining"),
    ("SDM", "SIAM International Conference on Data Mining", "B", "数据库/数据挖掘", "Databases & Data Mining"),
    ("RecSys", "ACM Conference on Recommender Systems", "B", "数据库/数据挖掘", "Databases & Data Mining"),
    ("WISE", "Web Information Systems Engineering Conference", "B", "数据库/数据挖掘", "Databases & Data Mining"),
    # === 数据库 C (13) ===
    ("APWeb", "Asia Pacific Web Conference", "C", "数据库/数据挖掘", "Databases & Data Mining"),
    ("DEXA", "International Conference on Database and Expert System Applications", "C", "数据库/数据挖掘", "Databases & Data Mining"),
    ("ECIR", "European Conference on Information Retrieval", "C", "数据库/数据挖掘", "Databases & Data Mining"),
    ("ESWC", "Extended Semantic Web Conference", "C", "数据库/数据挖掘", "Databases & Data Mining"),
    ("WebDB", "International Workshop on Web and Databases", "C", "数据库/数据挖掘", "Databases & Data Mining"),
    ("ER", "International Conference on Conceptual Modeling", "C", "数据库/数据挖掘", "Databases & Data Mining"),
    ("MDM", "International Conference on Mobile Data Management", "C", "数据库/数据挖掘", "Databases & Data Mining"),
    ("SSDBM", "International Conference on Scientific and Statistical Database Management", "C", "数据库/数据挖掘", "Databases & Data Mining"),
    ("WAIM", "International Conference on Web Age Information Management", "C", "数据库/数据挖掘", "Databases & Data Mining"),
    ("SSTD", "International Symposium on Spatial and Temporal Databases", "C", "数据库/数据挖掘", "Databases & Data Mining"),
    ("PAKDD", "Pacific-Asia Conference on Knowledge Discovery and Data Mining", "C", "数据库/数据挖掘", "Databases & Data Mining"),
    ("ADMA", "International Conference on Advanced Data Mining and Applications", "C", "数据库/数据挖掘", "Databases & Data Mining"),
    ("WISA", "Web Information Systems and Applications", "C", "数据库/数据挖掘", "Databases & Data Mining"),

    # === 计算机科学理论 A (5) ===
    ("STOC", "ACM Symposium on the Theory of Computing", "A", "计算机科学理论", "Theory of Computer Science"),
    ("SODA", "ACM-SIAM Symposium on Discrete Algorithms", "A", "计算机科学理论", "Theory of Computer Science"),
    ("CAV", "International Conference on Computer Aided Verification", "A", "计算机科学理论", "Theory of Computer Science"),
    ("FOCS", "IEEE Annual Symposium on Foundations of Computer Science", "A", "计算机科学理论", "Theory of Computer Science"),
    ("LICS", "ACM/IEEE Symposium on Logic in Computer Science", "A", "计算机科学理论", "Theory of Computer Science"),
    # === 理论 B (10) ===
    ("SoCG", "International Symposium on Computational Geometry", "B", "计算机科学理论", "Theory of Computer Science"),
    ("ESA", "European Symposium on Algorithms", "B", "计算机科学理论", "Theory of Computer Science"),
    ("CCC", "Conference on Computational Complexity", "B", "计算机科学理论", "Theory of Computer Science"),
    ("ICALP", "International Colloquium on Automata, Languages and Programming", "B", "计算机科学理论", "Theory of Computer Science"),
    ("CADE", "Conference on Automated Deduction", "B", "计算机科学理论", "Theory of Computer Science"),
    ("CONCUR", "International Conference on Concurrency Theory", "B", "计算机科学理论", "Theory of Computer Science"),
    ("HSCC", "International Conference on Hybrid Systems: Computation and Control", "B", "计算机科学理论", "Theory of Computer Science"),
    ("SAT", "International Conference on Theory and Applications of Satisfiability Testing", "B", "计算机科学理论", "Theory of Computer Science"),
    ("COCOON", "International Computing and Combinatorics Conference", "B", "计算机科学理论", "Theory of Computer Science"),
    ("FMCAD", "Formal Methods in Computer-Aided Design", "B", "计算机科学理论", "Theory of Computer Science"),
    # === 理论 C (10) ===
    ("CSL", "Computer Science Logic", "C", "计算机科学理论", "Theory of Computer Science"),
    ("FSTTCS", "Foundations of Software Technology and Theoretical Computer Science", "C", "计算机科学理论", "Theory of Computer Science"),
    ("DSAA", "IEEE International Conference on Data Science and Advanced Analytics", "C", "计算机科学理论", "Theory of Computer Science"),
    ("ICTAC", "International Colloquium on Theoretical Aspects of Computing", "C", "计算机科学理论", "Theory of Computer Science"),
    ("IPCO", "International Conference on Integer Programming and Combinatorial Optimization", "C", "计算机科学理论", "Theory of Computer Science"),
    ("FSCD", "International Conference on Formal Structures for Computation and Deduction", "C", "计算机科学理论", "Theory of Computer Science"),
    ("ISAAC", "International Symposium on Algorithms and Computation", "C", "计算机科学理论", "Theory of Computer Science"),
    ("MFCS", "International Conference on Mathematical Foundations of Computer Science", "C", "计算机科学理论", "Theory of Computer Science"),
    ("STACS", "Symposium on Theoretical Aspects of Computer Science", "C", "计算机科学理论", "Theory of Computer Science"),
    ("SETTA", "International Symposium on Software Engineering: Theories, Tools, and Applications", "C", "计算机科学理论", "Theory of Computer Science"),

    # === 图形学多媒体 A (4) ===
    ("ACM MM", "ACM International Conference on Multimedia", "A", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("SIGGRAPH", "ACM Special Interest Group on Computer Graphics", "A", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("VR", "IEEE Virtual Reality", "A", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("IEEE VIS", "IEEE Visualization Conference", "A", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    # === 图形学多媒体 B (14) ===
    ("ICMR", "ACM SIGMM International Conference on Multimedia Retrieval", "B", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("I3D", "ACM SIGGRAPH Symposium on Interactive 3D Graphics and Games", "B", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("SCA", "ACM SIGGRAPH/Eurographics Symposium on Computer Animation", "B", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("DCC", "Data Compression Conference", "B", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("Eurographics", "Annual Conference of the European Association for Computer Graphics", "B", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("EuroVis", "Eurographics Conference on Visualization", "B", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("SGP", "Eurographics Symposium on Geometry Processing", "B", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("EGSR", "Eurographics Symposium on Rendering", "B", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("ICASSP", "IEEE International Conference on Acoustics, Speech and Signal Processing", "B", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("ICME", "IEEE International Conference on Multimedia & Expo", "B", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("ISMAR", "International Symposium on Mixed and Augmented Reality", "B", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("PG", "Pacific Conference on Computer Graphics and Applications", "B", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("SPM", "Symposium on Solid and Physical Modeling", "B", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("INTERSPEECH", "Conference of the International Speech Communication Association", "B", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    # === 图形学多媒体 C (17) ===
    ("VRST", "ACM Symposium on Virtual Reality Software and Technology", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("CASA", "International Conference on Computer Animation and Social Agents", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("CGI", "Computer Graphics International", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("GMP", "Geometric Modeling and Processing", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("PacificVis", "IEEE Pacific Visualization Symposium", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("3DV", "International Conference on 3D Vision", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("CAD/Graphics", "International Conference on Computer-Aided Design and Computer Graphics", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("ICIP", "IEEE International Conference on Image Processing", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("MMM", "International Conference on Multimedia Modeling", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("MMAsia", "ACM Multimedia Asia", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("SMI", "Shape Modeling International", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("CVM", "Computational Visual Media", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("PRCV", "Chinese Conference on Pattern Recognition and Computer Vision", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("ICIG", "International Conference on Image and Graphics", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("NCMMSC", "National Conference on Man-Machine Speech Communication", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("ASRU", "Automatic Speech Recognition and Understanding Workshop", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
    ("SLT", "Spoken Language Technology", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),

    # === 人工智能 A (7) ===
    ("AAAI", "AAAI Conference on Artificial Intelligence", "A", "人工智能", "Artificial Intelligence"),
    ("NeurIPS", "Conference on Neural Information Processing Systems", "A", "人工智能", "Artificial Intelligence"),
    ("ACL", "Annual Meeting of the Association for Computational Linguistics", "A", "人工智能", "Artificial Intelligence"),
    ("CVPR", "IEEE/CVF Computer Vision and Pattern Recognition Conference", "A", "人工智能", "Artificial Intelligence"),
    ("ICCV", "International Conference on Computer Vision", "A", "人工智能", "Artificial Intelligence"),
    ("ICML", "International Conference on Machine Learning", "A", "人工智能", "Artificial Intelligence"),
    ("ICLR", "International Conference on Learning Representations", "A", "人工智能", "Artificial Intelligence"),
    # === 人工智能 B (14) ===
    ("COLT", "Annual Conference on Computational Learning Theory", "B", "人工智能", "Artificial Intelligence"),
    ("EMNLP", "Conference on Empirical Methods in Natural Language Processing", "B", "人工智能", "Artificial Intelligence"),
    ("ECAI", "European Conference on Artificial Intelligence", "B", "人工智能", "Artificial Intelligence"),
    ("ECCV", "European Conference on Computer Vision", "B", "人工智能", "Artificial Intelligence"),
    ("ICRA", "IEEE International Conference on Robotics and Automation", "B", "人工智能", "Artificial Intelligence"),
    ("ICAPS", "International Conference on Automated Planning and Scheduling", "B", "人工智能", "Artificial Intelligence"),
    ("ICCBR", "International Conference on Case-Based Reasoning", "B", "人工智能", "Artificial Intelligence"),
    ("COLING", "International Conference on Computational Linguistics", "B", "人工智能", "Artificial Intelligence"),
    ("KR", "International Conference on Principles of Knowledge Representation and Reasoning", "B", "人工智能", "Artificial Intelligence"),
    ("UAI", "Conference on Uncertainty in Artificial Intelligence", "B", "人工智能", "Artificial Intelligence"),
    ("AAMAS", "International Joint Conference on Autonomous Agents and Multi-agent Systems", "B", "人工智能", "Artificial Intelligence"),
    ("PPSN", "Parallel Problem Solving from Nature", "B", "人工智能", "Artificial Intelligence"),
    ("NAACL", "North American Chapter of the Association for Computational Linguistics", "B", "人工智能", "Artificial Intelligence"),
    ("IJCAI", "International Joint Conference on Artificial Intelligence", "B", "人工智能", "Artificial Intelligence"),
    # === 人工智能 C (22) ===
    ("AISTATS", "International Conference on Artificial Intelligence and Statistics", "C", "人工智能", "Artificial Intelligence"),
    ("ACCV", "Asian Conference on Computer Vision", "C", "人工智能", "Artificial Intelligence"),
    ("ACML", "Asian Conference on Machine Learning", "C", "人工智能", "Artificial Intelligence"),
    ("BMVC", "British Machine Vision Conference", "C", "人工智能", "Artificial Intelligence"),
    ("NLPCC", "CCF International Conference on Natural Language Processing and Chinese Computing", "C", "人工智能", "Artificial Intelligence"),
    ("CoNLL", "Conference on Computational Natural Language Learning", "C", "人工智能", "Artificial Intelligence"),
    ("GECCO", "Genetic and Evolutionary Computation Conference", "C", "人工智能", "Artificial Intelligence"),
    ("ICTAI", "IEEE International Conference on Tools with Artificial Intelligence", "C", "人工智能", "Artificial Intelligence"),
    ("IROS", "IEEE/RSJ International Conference on Intelligent Robots and Systems", "C", "人工智能", "Artificial Intelligence"),
    ("ALT", "International Conference on Algorithmic Learning Theory", "C", "人工智能", "Artificial Intelligence"),
    ("ICANN", "International Conference on Artificial Neural Networks", "C", "人工智能", "Artificial Intelligence"),
    ("FG", "IEEE International Conference on Automatic Face and Gesture Recognition", "C", "人工智能", "Artificial Intelligence"),
    ("ICDAR", "International Conference on Document Analysis and Recognition", "C", "人工智能", "Artificial Intelligence"),
    ("ILP", "International Conference on Inductive Logic Programming", "C", "人工智能", "Artificial Intelligence"),
    ("KSEM", "International Conference on Knowledge Science, Engineering and Management", "C", "人工智能", "Artificial Intelligence"),
    ("ICONIP", "International Conference on Neural Information Processing", "C", "人工智能", "Artificial Intelligence"),
    ("ICPR", "International Conference on Pattern Recognition", "C", "人工智能", "Artificial Intelligence"),
    ("IJCB", "International Joint Conference on Biometrics", "C", "人工智能", "Artificial Intelligence"),
    ("IJCNN", "International Joint Conference on Neural Networks", "C", "人工智能", "Artificial Intelligence"),
    ("PRICAI", "Pacific Rim International Conference on Artificial Intelligence", "C", "人工智能", "Artificial Intelligence"),
    ("IEEE CEC", "Congress on Evolutionary Computation", "C", "人工智能", "Artificial Intelligence"),
    ("DAI", "International Conference on Distributed Artificial Intelligence", "C", "人工智能", "Artificial Intelligence"),

    # === 人机交互 A (4) ===
    ("CSCW", "ACM Conference On Computer-Supported Cooperative Work And Social Computing", "A", "人机交互与普适计算", "HCI & Pervasive Computing"),
    ("CHI", "ACM Conference on Human Factors in Computing Systems", "A", "人机交互与普适计算", "HCI & Pervasive Computing"),
    ("UbiComp", "ACM International Joint Conference on Pervasive and Ubiquitous Computing", "A", "人机交互与普适计算", "HCI & Pervasive Computing"),
    ("UIST", "ACM Symposium on User Interface Software and Technology", "A", "人机交互与普适计算", "HCI & Pervasive Computing"),
    # === 人机交互 B (7) ===
    ("GROUP", "ACM International Conference on Supporting Group Work", "B", "人机交互与普适计算", "HCI & Pervasive Computing"),
    ("IUI", "ACM International Conference on Intelligent User Interfaces", "B", "人机交互与普适计算", "HCI & Pervasive Computing"),
    ("ISS", "ACM International Conference on Interactive Surfaces and Spaces", "B", "人机交互与普适计算", "HCI & Pervasive Computing"),
    ("ECSCW", "European Conference on Computer Supported Cooperative Work", "B", "人机交互与普适计算", "HCI & Pervasive Computing"),
    ("PERCOM", "IEEE International Conference on Pervasive Computing and Communications", "B", "人机交互与普适计算", "HCI & Pervasive Computing"),
    ("MobileHCI", "ACM International Conference on Mobile Human-Computer Interaction", "B", "人机交互与普适计算", "HCI & Pervasive Computing"),
    ("ICWSM", "International AAAI Conference on Web and Social Media", "B", "人机交互与普适计算", "HCI & Pervasive Computing"),
    # === 人机交互 C (15) ===
    ("DIS", "ACM SIGCHI Conference on Designing Interactive Systems", "C", "人机交互与普适计算", "HCI & Pervasive Computing"),
    ("ICMI", "ACM International Conference on Multimodal Interaction", "C", "人机交互与普适计算", "HCI & Pervasive Computing"),
    ("ASSETS", "International ACM SIGACCESS Conference on Computers and Accessibility", "C", "人机交互与普适计算", "HCI & Pervasive Computing"),
    ("GI", "Graphics Interface", "C", "人机交互与普适计算", "HCI & Pervasive Computing"),
    ("UIC", "IEEE International Conference on Ubiquitous Intelligence and Computing", "C", "人机交互与普适计算", "HCI & Pervasive Computing"),
    ("World Haptics", "IEEE World Haptics Conference", "C", "人机交互与普适计算", "HCI & Pervasive Computing"),
    ("INTERACT", "International Conference on Human-Computer Interaction of International Federation for Information Processing", "C", "人机交互与普适计算", "HCI & Pervasive Computing"),
    ("IDC", "ACM Interaction Design and Children", "C", "人机交互与普适计算", "HCI & Pervasive Computing"),
    ("CollaborateCom", "International Conference on Collaborative Computing: Networking, Applications and Worksharing", "C", "人机交互与普适计算", "HCI & Pervasive Computing"),
    ("CSCWD", "International Conference on Computer Supported Cooperative Work in Design", "C", "人机交互与普适计算", "HCI & Pervasive Computing"),
    ("CoopIS", "International Conference on Cooperative Information Systems", "C", "人机交互与普适计算", "HCI & Pervasive Computing"),
    ("MobiQuitous", "International Conference on Mobile and Ubiquitous Systems: Computing, Networking and Services", "C", "人机交互与普适计算", "HCI & Pervasive Computing"),
    ("AVI", "International Working Conference on Advanced Visual Interfaces", "C", "人机交互与普适计算", "HCI & Pervasive Computing"),
    ("GPC", "Conference on Green, Pervasive and Cloud Computing", "C", "人机交互与普适计算", "HCI & Pervasive Computing"),
    ("ICXR", "CCF International Conference on Extended Reality", "C", "人机交互与普适计算", "HCI & Pervasive Computing"),

    # === 交叉综合 A (2) ===
    ("WWW", "International World Wide Web Conference", "A", "交叉/综合/新兴", "Cross-disciplinary"),
    ("RTSS", "IEEE Real-Time Systems Symposium", "A", "交叉/综合/新兴", "Cross-disciplinary"),
    # === 交叉综合 B (7) ===
    ("CogSci", "Annual Meeting of the Cognitive Science Society", "B", "交叉/综合/新兴", "Cross-disciplinary"),
    ("BIBM", "IEEE International Conference on Bioinformatics and Biomedicine", "B", "交叉/综合/新兴", "Cross-disciplinary"),
    ("EMSOFT", "International Conference on Embedded Software", "B", "交叉/综合/新兴", "Cross-disciplinary"),
    ("ISMB", "International Conference on Intelligent Systems for Molecular Biology", "B", "交叉/综合/新兴", "Cross-disciplinary"),
    ("RECOMB", "Annual International Conference on Research in Computational Molecular Biology", "B", "交叉/综合/新兴", "Cross-disciplinary"),
    ("MICCAI", "International Conference on Medical Image Computing and Computer-Assisted Intervention", "B", "交叉/综合/新兴", "Cross-disciplinary"),
    ("WINE", "Conference on Web and Internet Economics", "B", "交叉/综合/新兴", "Cross-disciplinary"),
    # === 交叉综合 C (13) ===
    ("AMIA", "American Medical Informatics Association Annual Symposium", "C", "交叉/综合/新兴", "Cross-disciplinary"),
    ("APBC", "Asia Pacific Bioinformatics Conference", "C", "交叉/综合/新兴", "Cross-disciplinary"),
    ("IEEE BigData", "IEEE International Conference on Big Data", "C", "交叉/综合/新兴", "Cross-disciplinary"),
    ("IEEE CLOUD", "IEEE International Conference on Cloud Computing", "C", "交叉/综合/新兴", "Cross-disciplinary"),
    ("SMC", "IEEE International Conference on Systems, Man, and Cybernetics", "C", "交叉/综合/新兴", "Cross-disciplinary"),
    ("COSIT", "International Conference on Spatial Information Theory", "C", "交叉/综合/新兴", "Cross-disciplinary"),
    ("ISBRA", "International Symposium on Bioinformatics Research and Applications", "C", "交叉/综合/新兴", "Cross-disciplinary"),
    ("SAGT", "International Symposium on Algorithmic Game Theory", "C", "交叉/综合/新兴", "Cross-disciplinary"),
    ("SIGSPATIAL", "ACM Special Interest Group on Spatial Information", "C", "交叉/综合/新兴", "Cross-disciplinary"),
    ("ICIC", "International Conference on Intelligent Computing", "C", "交叉/综合/新兴", "Cross-disciplinary"),
    ("ICSS", "International Conference on Service Science", "C", "交叉/综合/新兴", "Cross-disciplinary"),
    ("AFT", "Advances in Financial Technologies", "C", "交叉/综合/新兴", "Cross-disciplinary"),
    ("IJTCS-FAW", "International Joint Conference on Theoretical Computer Science - Frontier of Algorithmic Wisdom", "C", "交叉/综合/新兴", "Cross-disciplinary"),
]

# Journal-type conferences
JOURNAL_TYPE = {"CHES", "TCHES", "FSE"}

# ============================================================
def normalize(s):
    if not s: return ""
    s = s.lower().strip().replace('\n',' ').replace('\r',' ').replace('\xa0',' ')
    s = re.sub(r'\s+', ' ', s).replace('&', 'and')
    return s

def generate_id(abbr):
    return re.sub(r'[^a-z0-9-]', '', abbr.lower().replace('&','-').replace(' ','-').replace('/','-'))

def slugify(text):
    return re.sub(r'[^a-z0-9]+', '-', text.lower()).strip('-')

def load_excel(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    return [row for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)]

def match_entry(full_name, official_list):
    """Match using full official names with SequenceMatcher."""
    name_norm = normalize(full_name)
    best_score, best_match = 0, None

    for abbr, full_official, rank, cat_zh, cat_en in official_list:
        off_norm = normalize(full_official)
        score = SequenceMatcher(None, name_norm, off_norm).ratio()
        if score > best_score:
            best_score = score
            best_match = (abbr, rank, cat_zh, cat_en)

    return (best_match, best_score) if best_score >= 0.80 else (None, best_score)

def build_conferences():
    raw = load_excel(BASE_DIR / "conference_map.xlsx")
    results = []
    unmatched = []

    for row in raw:
        name = str(row[1]).strip() if row[1] else ""
        if not name: continue

        # Fixes
        if "ACM SIGOPS Annual Technical Conference" in name:
            name = "USENIX Annual Technical Conference"
        if "Virtual Reality and Visualization" in name and "International Conference on Virtual Reality" in name:
            print(f"  SKIP (not in 2026 CCF): {name}")
            continue

        match, score = match_entry(name, CCF_CONF_FULL)
        if match:
            abbr, rank, cat_zh, cat_en = match
            results.append(OrderedDict([
                ("id", generate_id(abbr)),
                ("abbreviation", abbr),
                ("full_name", name),
                ("type", "conference"),
                ("sub_type", "journal-type" if abbr in JOURNAL_TYPE else None),
                ("ccf_rank", rank),
                ("category_zh", cat_zh),
                ("category_en", cat_en),
                ("aliases", []),
            ]))
        else:
            unmatched.append((name, score))

    rank_order = {"A": 0, "B": 1, "C": 2}
    results.sort(key=lambda c: (rank_order[c["ccf_rank"]], c["abbreviation"]))

    print(f"Conferences: {len(results)} matched, {len(unmatched)} unmatched")
    for name, score in unmatched:
        print(f"  UNMATCHED [{score:.2f}]: {name[:100]}")
    return results

def build_journals():
    # For journals, use the original keyword-based approach since it matched 297/297
    # Import from the previous working version
    raw = load_excel(BASE_DIR / "journal_map.xlsx")
    results = []
    seen_ids = set()

    official_journals = [
        ("TOCS", "ACM Transactions on Computer Systems", "A", "计算机体系结构", "Computer Architecture"),
        ("TOS", "ACM Transactions on Storage", "A", "计算机体系结构", "Computer Architecture"),
        ("TCAD", "IEEE Transactions on Computer-Aided Design of Integrated Circuits and Systems", "A", "计算机体系结构", "Computer Architecture"),
        ("TC", "IEEE Transactions on Computers", "A", "计算机体系结构", "Computer Architecture"),
        ("TPDS", "IEEE Transactions on Parallel and Distributed Systems", "A", "计算机体系结构", "Computer Architecture"),
        ("TACO", "ACM Transactions on Architecture and Code Optimization", "A", "计算机体系结构", "Computer Architecture"),
        ("TAAS", "ACM Transactions on Autonomous and Adaptive Systems", "B", "计算机体系结构", "Computer Architecture"),
        ("TODAES", "ACM Transactions on Design Automation of Electronic Systems", "B", "计算机体系结构", "Computer Architecture"),
        ("TECS", "ACM Transactions on Embedded Computing Systems", "B", "计算机体系结构", "Computer Architecture"),
        ("TRETS", "ACM Transactions on Reconfigurable Technology and Systems", "B", "计算机体系结构", "Computer Architecture"),
        ("TVLSI", "IEEE Transactions on Very Large Scale Integration (VLSI) Systems", "B", "计算机体系结构", "Computer Architecture"),
        ("JPDC", "Journal of Parallel and Distributed Computing", "B", "计算机体系结构", "Computer Architecture"),
        ("JSA", "Journal of Systems Architecture: Embedded Software Design", "B", "计算机体系结构", "Computer Architecture"),
        ("Parallel Computing", "Parallel Computing", "B", "计算机体系结构", "Computer Architecture"),
        ("Performance Evaluation", "Performance Evaluation: An International Journal", "B", "计算机体系结构", "Computer Architecture"),
        ("TCC", "IEEE Transactions on Cloud Computing", "B", "计算机体系结构", "Computer Architecture"),
        ("JETC", "ACM Journal on Emerging Technologies in Computing Systems", "C", "计算机体系结构", "Computer Architecture"),
        ("CCPE", "Concurrency and Computation: Practice and Experience", "C", "计算机体系结构", "Computer Architecture"),
        ("DC", "Distributed Computing", "C", "计算机体系结构", "Computer Architecture"),
        ("FGCS", "Future Generation Computer Systems", "C", "计算机体系结构", "Computer Architecture"),
        ("Integration", "Integration, the VLSI Journal", "C", "计算机体系结构", "Computer Architecture"),
        ("JETTA", "Journal of Electronic Testing-Theory and Applications", "C", "计算机体系结构", "Computer Architecture"),
        ("JGC", "Journal of Grid computing", "C", "计算机体系结构", "Computer Architecture"),
        ("RTS", "Real-Time Systems", "C", "计算机体系结构", "Computer Architecture"),
        ("TJSC", "The Journal of Supercomputing", "C", "计算机体系结构", "Computer Architecture"),
        ("TCASI", "IEEE Transactions on Circuits and Systems I: Regular Papers", "C", "计算机体系结构", "Computer Architecture"),
        ("CCF-THPC", "CCF Transactions on High Performance Computing", "C", "计算机体系结构", "Computer Architecture"),
        ("TSUSC", "IEEE Transactions on Sustainable Computing", "C", "计算机体系结构", "Computer Architecture"),
        ("JSAC", "IEEE Journal on Selected Areas in Communications", "A", "计算机网络", "Computer Networks"),
        ("TMC", "IEEE Transactions on Mobile Computing", "A", "计算机网络", "Computer Networks"),
        ("TON", "IEEE/ACM Transactions on Networking", "A", "计算机网络", "Computer Networks"),
        ("TOIT", "ACM Transactions on Internet Technology", "B", "计算机网络", "Computer Networks"),
        ("TOMM", "ACM Transactions on Multimedia Computing, Communications and Applications", "B", "计算机网络", "Computer Networks"),
        ("TOSN", "ACM Transactions on Sensor Networks", "B", "计算机网络", "Computer Networks"),
        ("CN", "Computer Networks", "B", "计算机网络", "Computer Networks"),
        ("TCOM", "IEEE Transactions on Communications", "B", "计算机网络", "Computer Networks"),
        ("TWC", "IEEE Transactions on Wireless Communications", "B", "计算机网络", "Computer Networks"),
        ("Ad Hoc Networks", "Ad Hoc Networks", "C", "计算机网络", "Computer Networks"),
        ("CC", "Computer Communications", "C", "计算机网络", "Computer Networks"),
        ("TNSM", "IEEE Transactions on Network and Service Management", "C", "计算机网络", "Computer Networks"),
        ("IET Communications", "IET Communications", "C", "计算机网络", "Computer Networks"),
        ("JNCA", "Journal of Network and Computer Applications", "C", "计算机网络", "Computer Networks"),
        ("MONET", "Mobile Networks and Applications", "C", "计算机网络", "Computer Networks"),
        ("Networks", "Networks", "C", "计算机网络", "Computer Networks"),
        ("PPNA", "Peer-to-Peer Networking and Applications", "C", "计算机网络", "Computer Networks"),
        ("WCMC", "Wireless Communications and Mobile Computing", "C", "计算机网络", "Computer Networks"),
        ("Wireless Networks", "Wireless Networks", "C", "计算机网络", "Computer Networks"),
        ("IOT", "IEEE Internet of Things Journal", "C", "计算机网络", "Computer Networks"),
        ("TIOT", "ACM Transactions on Internet of Things", "C", "计算机网络", "Computer Networks"),
        ("TDSC", "IEEE Transactions on Dependable and Secure Computing", "A", "网络与信息安全", "Network & Information Security"),
        ("TIFS", "IEEE Transactions on Information Forensics and Security", "A", "网络与信息安全", "Network & Information Security"),
        ("Journal of Cryptology", "Journal of Cryptology", "A", "网络与信息安全", "Network & Information Security"),
        ("TOPS", "ACM Transactions on Privacy and Security", "B", "网络与信息安全", "Network & Information Security"),
        ("Computers & Security", "Computers & Security", "B", "网络与信息安全", "Network & Information Security"),
        ("Designs Codes and Cryptography", "Designs, Codes and Cryptography", "B", "网络与信息安全", "Network & Information Security"),
        ("JCS", "Journal of Computer Security", "B", "网络与信息安全", "Network & Information Security"),
        ("Cybersecurity", "Cybersecurity", "B", "网络与信息安全", "Network & Information Security"),
        ("CLSR", "Computer Law & Security Review", "C", "网络与信息安全", "Network & Information Security"),
        ("EURASIP JIS", "EURASIP Journal on Information Security", "C", "网络与信息安全", "Network & Information Security"),
        ("IET Information Security", "IET Information Security", "C", "网络与信息安全", "Network & Information Security"),
        ("IMCS", "Information and Computer Security", "C", "网络与信息安全", "Network & Information Security"),
        ("IJICS", "International Journal of Information and Computer Security", "C", "网络与信息安全", "Network & Information Security"),
        ("IJISP", "International Journal of Information Security and Privacy", "C", "网络与信息安全", "Network & Information Security"),
        ("JISA", "Journal of Information Security and Applications", "C", "网络与信息安全", "Network & Information Security"),
        ("SCN", "Security and Communication Networks", "C", "网络与信息安全", "Network & Information Security"),
        ("HCC", "High-Confidence Computing", "C", "网络与信息安全", "Network & Information Security"),
        ("TOPLAS", "ACM Transactions on Programming Languages and Systems", "A", "软件工程", "Software Engineering"),
        ("TOSEM", "ACM Transactions on Software Engineering and Methodology", "A", "软件工程", "Software Engineering"),
        ("TSE", "IEEE Transactions on Software Engineering", "A", "软件工程", "Software Engineering"),
        ("TSC", "IEEE Transactions on Services Computing", "A", "软件工程", "Software Engineering"),
        ("ASE J", "Automated Software Engineering", "B", "软件工程", "Software Engineering"),
        ("ESE", "Empirical Software Engineering", "B", "软件工程", "Software Engineering"),
        ("IETS", "IET Software", "B", "软件工程", "Software Engineering"),
        ("IST", "Information and Software Technology", "B", "软件工程", "Software Engineering"),
        ("JFP", "Journal of Functional Programming", "B", "软件工程", "Software Engineering"),
        ("JSME", "Journal of Software: Evolution and Process", "B", "软件工程", "Software Engineering"),
        ("JSS", "Journal of Systems and Software", "B", "软件工程", "Software Engineering"),
        ("RE J", "Requirements Engineering", "B", "软件工程", "Software Engineering"),
        ("SCP", "Science of Computer Programming", "B", "软件工程", "Software Engineering"),
        ("SoSyM", "Software and Systems Modeling", "B", "软件工程", "Software Engineering"),
        ("STVR", "Software Testing, Verification and Reliability", "B", "软件工程", "Software Engineering"),
        ("SPE", "Software: Practice and Experience", "B", "软件工程", "Software Engineering"),
        ("CL", "Computer Languages, Systems and Structures", "C", "软件工程", "Software Engineering"),
        ("IJSEKE", "International Journal of Software Engineering and Knowledge Engineering", "C", "软件工程", "Software Engineering"),
        ("STTT", "International Journal of Software Tools for Technology Transfer", "C", "软件工程", "Software Engineering"),
        ("JLAMP", "Journal of Logical and Algebraic Methods in Programming", "C", "软件工程", "Software Engineering"),
        ("JWE", "Journal of Web Engineering", "C", "软件工程", "Software Engineering"),
        ("SOCA", "Service Oriented Computing and Applications", "C", "软件工程", "Software Engineering"),
        ("SQJ", "Software Quality Journal", "C", "软件工程", "Software Engineering"),
        ("TPLP", "Theory and Practice of Logic Programming", "C", "软件工程", "Software Engineering"),
        ("PACMPL", "Proceedings of the ACM on Programming Languages", "C", "软件工程", "Software Engineering"),
        ("TODS", "ACM Transactions on Database Systems", "A", "数据库/数据挖掘", "Databases & Data Mining"),
        ("TOIS", "ACM Transactions on Information Systems", "A", "数据库/数据挖掘", "Databases & Data Mining"),
        ("TKDE", "IEEE Transactions on Knowledge and Data Engineering", "A", "数据库/数据挖掘", "Databases & Data Mining"),
        ("VLDBJ", "The VLDB Journal", "A", "数据库/数据挖掘", "Databases & Data Mining"),
        ("TKDD", "ACM Transactions on Knowledge Discovery from Data", "B", "数据库/数据挖掘", "Databases & Data Mining"),
        ("TWEB", "ACM Transactions on the Web", "B", "数据库/数据挖掘", "Databases & Data Mining"),
        ("AEI", "Advanced Engineering Informatics", "B", "数据库/数据挖掘", "Databases & Data Mining"),
        ("DKE", "Data & Knowledge Engineering", "B", "数据库/数据挖掘", "Databases & Data Mining"),
        ("DMKD", "Data Mining and Knowledge Discovery", "B", "数据库/数据挖掘", "Databases & Data Mining"),
        ("EJIS", "European Journal of Information Systems", "B", "数据库/数据挖掘", "Databases & Data Mining"),
        ("GeoInformatica", "GeoInformatica", "B", "数据库/数据挖掘", "Databases & Data Mining"),
        ("IPM", "Information Processing and Management", "B", "数据库/数据挖掘", "Databases & Data Mining"),
        ("Information Sciences", "Information Sciences", "B", "数据库/数据挖掘", "Databases & Data Mining"),
        ("IS", "Information Systems", "B", "数据库/数据挖掘", "Databases & Data Mining"),
        ("JASIST", "Journal of the Association for Information Science and Technology", "B", "数据库/数据挖掘", "Databases & Data Mining"),
        ("JWS", "Journal of Web Semantics", "B", "数据库/数据挖掘", "Databases & Data Mining"),
        ("KAIS", "Knowledge and Information Systems", "B", "数据库/数据挖掘", "Databases & Data Mining"),
        ("DSE", "Data Science and Engineering", "B", "数据库/数据挖掘", "Databases & Data Mining"),
        ("DPD", "Distributed and Parallel Databases", "C", "数据库/数据挖掘", "Databases & Data Mining"),
        ("I&M", "Information & Management", "C", "数据库/数据挖掘", "Databases & Data Mining"),
        ("IPL", "Information Processing Letters", "C", "数据库/数据挖掘", "Databases & Data Mining"),
        ("IRJ", "Information Retrieval Journal", "C", "数据库/数据挖掘", "Databases & Data Mining"),
        ("IJCIS", "International Journal of Cooperative Information Systems", "C", "数据库/数据挖掘", "Databases & Data Mining"),
        ("IJGIS", "International Journal of Geographical Information Science", "C", "数据库/数据挖掘", "Databases & Data Mining"),
        ("IJIS", "International Journal of Intelligent Systems", "C", "数据库/数据挖掘", "Databases & Data Mining"),
        ("IJKM", "International Journal of Knowledge Management", "C", "数据库/数据挖掘", "Databases & Data Mining"),
        ("IJSWIS", "International Journal on Semantic Web and Information Systems", "C", "数据库/数据挖掘", "Databases & Data Mining"),
        ("JCIS", "Journal of Computer Information Systems", "C", "数据库/数据挖掘", "Databases & Data Mining"),
        ("JDM", "Journal of Database Management", "C", "数据库/数据挖掘", "Databases & Data Mining"),
        ("JGITM", "Journal of Global Information Technology Management", "C", "数据库/数据挖掘", "Databases & Data Mining"),
        ("JIIS", "Journal of Intelligent Information Systems", "C", "数据库/数据挖掘", "Databases & Data Mining"),
        ("JSIS", "The Journal of Strategic Information Systems", "C", "数据库/数据挖掘", "Databases & Data Mining"),
        ("TIST", "ACM Transactions on Intelligent Systems and Technology", "C", "数据库/数据挖掘", "Databases & Data Mining"),
        ("TORS", "ACM Transactions on Recommender Systems", "C", "数据库/数据挖掘", "Databases & Data Mining"),
        ("TIT", "IEEE Transactions on Information Theory", "A", "计算机科学理论", "Theory of Computer Science"),
        ("IANDC", "Information and Computation", "A", "计算机科学理论", "Theory of Computer Science"),
        ("SICOMP", "SIAM Journal on Computing", "A", "计算机科学理论", "Theory of Computer Science"),
        ("TALG", "ACM Transactions on Algorithms", "B", "计算机科学理论", "Theory of Computer Science"),
        ("TOCL", "ACM Transactions on Computational Logic", "B", "计算机科学理论", "Theory of Computer Science"),
        ("TOMS", "ACM Transactions on Mathematical Software", "B", "计算机科学理论", "Theory of Computer Science"),
        ("Algorithmica", "Algorithmica", "B", "计算机科学理论", "Theory of Computer Science"),
        ("CC J", "Computational Complexity", "B", "计算机科学理论", "Theory of Computer Science"),
        ("FAC", "Formal Aspects of Computing", "B", "计算机科学理论", "Theory of Computer Science"),
        ("FMSD", "Formal Methods in System Design", "B", "计算机科学理论", "Theory of Computer Science"),
        ("INFORMS", "INFORMS Journal on Computing", "B", "计算机科学理论", "Theory of Computer Science"),
        ("JCSS", "Journal of Computer and System Sciences", "B", "计算机科学理论", "Theory of Computer Science"),
        ("JGO", "Journal of Global Optimization", "B", "计算机科学理论", "Theory of Computer Science"),
        ("JSC", "Journal of Symbolic Computation", "B", "计算机科学理论", "Theory of Computer Science"),
        ("MSCS", "Mathematical Structures in Computer Science", "B", "计算机科学理论", "Theory of Computer Science"),
        ("TCS", "Theoretical Computer Science", "B", "计算机科学理论", "Theory of Computer Science"),
        ("ACTA", "Acta Informatica", "C", "计算机科学理论", "Theory of Computer Science"),
        ("APAL", "Annals of Pure and Applied Logic", "C", "计算机科学理论", "Theory of Computer Science"),
        ("DAM", "Discrete Applied Mathematics", "C", "计算机科学理论", "Theory of Computer Science"),
        ("FUIN", "Fundamenta Informaticae", "C", "计算机科学理论", "Theory of Computer Science"),
        ("IPL T", "Information Processing Letters", "C", "计算机科学理论", "Theory of Computer Science"),
        ("JCOMPLEXITY", "Journal of Complexity", "C", "计算机科学理论", "Theory of Computer Science"),
        ("LOGCOM", "Journal of Logic and Computation", "C", "计算机科学理论", "Theory of Computer Science"),
        ("JSL", "The Journal of Symbolic Logic", "C", "计算机科学理论", "Theory of Computer Science"),
        ("LMCS", "Logical Methods in Computer Science", "C", "计算机科学理论", "Theory of Computer Science"),
        ("SIDMA", "SIAM Journal on Discrete Mathematics", "C", "计算机科学理论", "Theory of Computer Science"),
        ("TOCS T", "Theory of Computing Systems", "C", "计算机科学理论", "Theory of Computer Science"),
        ("TQC", "ACM Transactions in Quantum Computing", "C", "计算机科学理论", "Theory of Computer Science"),
        ("TOG", "ACM Transactions on Graphics", "A", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("TIP", "IEEE Transactions on Image Processing", "A", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("TVCG", "IEEE Transactions on Visualization and Computer Graphics", "A", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("TMM", "IEEE Transactions on Multimedia", "A", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("TOMM G", "ACM Transactions on Multimedia Computing, Communications and Applications", "B", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("CAGD", "Computer Aided Geometric Design", "B", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("CGF", "Computer Graphics Forum", "B", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("CAD", "Computer-Aided Design", "B", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("TCSVT", "IEEE Transactions on Circuits and Systems for Video Technology", "B", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("JASA", "The Journal of the Acoustical Society of America", "B", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("SIIMS", "SIAM Journal on Imaging Sciences", "B", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("SPECOM", "Speech Communication", "B", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("CVMJ", "Computational Visual Media", "B", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("CGTA", "Computational Geometry: Theory and Applications", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("CAVW", "Computer animation & virtual worlds", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("C&G", "Computers & Graphics", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("DCG", "Discrete & Computational Geometry", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("SPL", "IEEE Signal Processing Letters", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("IET-IPR", "IET Image Processing", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("JVCIR", "Journal of Visual Communication and Image  Representation", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("MS", "Multimedia Systems", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("MTA", "Multimedia Tools and Applications", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("SIGPRO", "Signal Processing", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("IMAGE", "Signal Processing: Image Communication", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("TVC", "The Visual Computer", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("VI", "Visual Informatics", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("VRIH", "Virtual Reality & Intelligent Hardware", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("GMOD", "Graphical Models", "C", "计算机图形学与多媒体", "Computer Graphics & Multimedia"),
        ("AI", "Artificial Intelligence", "A", "人工智能", "Artificial Intelligence"),
        ("TPAMI", "IEEE Transactions on Pattern Analysis and Machine Intelligence", "A", "人工智能", "Artificial Intelligence"),
        ("IJCV", "International Journal of Computer Vision", "A", "人工智能", "Artificial Intelligence"),
        ("JMLR", "Journal of Machine Learning Research", "A", "人工智能", "Artificial Intelligence"),
        ("TAP", "ACM Transactions on Applied Perception", "B", "人工智能", "Artificial Intelligence"),
        ("AAMAS J", "Autonomous Agents and Multi-Agent Systems", "B", "人工智能", "Artificial Intelligence"),
        ("Computational Linguistics", "Computational Linguistics", "B", "人工智能", "Artificial Intelligence"),
        ("CVIU", "Computer Vision and Image Understanding", "B", "人工智能", "Artificial Intelligence"),
        ("DKE AI", "Data & Knowledge Engineering", "B", "人工智能", "Artificial Intelligence"),
        ("Evolutionary Computation", "Evolutionary Computation", "B", "人工智能", "Artificial Intelligence"),
        ("TAC", "IEEE Transactions on Affective Computing", "B", "人工智能", "Artificial Intelligence"),
        ("TASLP", "IEEE/ACM Transactions on Audio, Speech and Language Processing", "B", "人工智能", "Artificial Intelligence"),
        ("TCYB", "IEEE Transactions on Cybernetics", "B", "人工智能", "Artificial Intelligence"),
        ("TEC", "IEEE Transactions on Evolutionary Computation", "B", "人工智能", "Artificial Intelligence"),
        ("TFS", "IEEE Transactions on Fuzzy Systems", "B", "人工智能", "Artificial Intelligence"),
        ("TNNLS", "IEEE Transactions on Neural Networks and learning systems", "B", "人工智能", "Artificial Intelligence"),
        ("IJAR", "International Journal of Approximate Reasoning", "B", "人工智能", "Artificial Intelligence"),
        ("JAIR", "Journal of Artificial Intelligence Research", "B", "人工智能", "Artificial Intelligence"),
        ("JAR", "Journal of Automated Reasoning", "B", "人工智能", "Artificial Intelligence"),
        ("JSLHR", "Journal of Speech, Language, and Hearing Research", "B", "人工智能", "Artificial Intelligence"),
        ("Machine Learning", "Machine Learning", "B", "人工智能", "Artificial Intelligence"),
        ("Neural Computation", "Neural Computation", "B", "人工智能", "Artificial Intelligence"),
        ("Neural Networks", "Neural Networks", "B", "人工智能", "Artificial Intelligence"),
        ("PR", "Pattern Recognition", "B", "人工智能", "Artificial Intelligence"),
        ("TACL", "Transactions of the Association for Computational Linguistics", "B", "人工智能", "Artificial Intelligence"),
        ("TALLIP", "ACM Transactions on Asian and Low-Resource Language Information Processing", "C", "人工智能", "Artificial Intelligence"),
        ("Applied Intelligence", "Applied Intelligence", "C", "人工智能", "Artificial Intelligence"),
        ("AIM", "Artificial Intelligence in Medicine", "C", "人工智能", "Artificial Intelligence"),
        ("Artificial Life", "Artificial Life", "C", "人工智能", "Artificial Intelligence"),
        ("Computational Intelligence", "Computational Intelligence", "C", "人工智能", "Artificial Intelligence"),
        ("Computer Speech & Language", "Computer Speech & Language", "C", "人工智能", "Artificial Intelligence"),
        ("Connection Science", "Connection Science", "C", "人工智能", "Artificial Intelligence"),
        ("DSS", "Decision Support Systems", "C", "人工智能", "Artificial Intelligence"),
        ("EAAI", "Engineering Applications of Artificial Intelligence", "C", "人工智能", "Artificial Intelligence"),
        ("ESWA", "Expert Systems with Applications", "C", "人工智能", "Artificial Intelligence"),
        ("Fuzzy Sets and Systems", "Fuzzy Sets and Systems", "C", "人工智能", "Artificial Intelligence"),
        ("TG", "IEEE Transactions on Games", "C", "人工智能", "Artificial Intelligence"),
        ("IET-CVI", "IET Computer Vision", "C", "人工智能", "Artificial Intelligence"),
        ("IET Signal Processing", "IET Signal Processing", "C", "人工智能", "Artificial Intelligence"),
        ("IVC", "Image and Vision Computing", "C", "人工智能", "Artificial Intelligence"),
        ("IDA", "Intelligent Data Analysis", "C", "人工智能", "Artificial Intelligence"),
        ("IJCIA", "International Journal of Computational Intelligence and Applications", "C", "人工智能", "Artificial Intelligence"),
        ("IJIS AI", "International Journal of Intelligent Systems", "C", "人工智能", "Artificial Intelligence"),
        ("IJNS", "International Journal of Neural Systems", "C", "人工智能", "Artificial Intelligence"),
        ("IJPRAI", "International Journal of Pattern Recognition and Artificial Intelligence", "C", "人工智能", "Artificial Intelligence"),
        ("IJUFKS", "International Journal of Uncertainty, Fuzziness and Knowledge-Based Systems", "C", "人工智能", "Artificial Intelligence"),
        ("IJDAR", "International Journal on Document Analysis and Recognition", "C", "人工智能", "Artificial Intelligence"),
        ("JETAI", "Journal of Experimental and Theoretical Artificial Intelligence", "C", "人工智能", "Artificial Intelligence"),
        ("KBS", "Knowledge-Based Systems", "C", "人工智能", "Artificial Intelligence"),
        ("Machine Translation", "Machine Translation", "C", "人工智能", "Artificial Intelligence"),
        ("Machine Vision and Applications", "Machine Vision and Applications", "C", "人工智能", "Artificial Intelligence"),
        ("Natural Computing", "Natural Computing", "C", "人工智能", "Artificial Intelligence"),
        ("NLE", "Natural Language Engineering", "C", "人工智能", "Artificial Intelligence"),
        ("NCA", "Neural Computing and Applications", "C", "人工智能", "Artificial Intelligence"),
        ("NPL", "Neural Processing Letters", "C", "人工智能", "Artificial Intelligence"),
        ("Neurocomputing", "Neurocomputing", "C", "人工智能", "Artificial Intelligence"),
        ("PAA", "Pattern Analysis and Applications", "C", "人工智能", "Artificial Intelligence"),
        ("PRL", "Pattern Recognition Letters", "C", "人工智能", "Artificial Intelligence"),
        ("Soft Computing", "Soft Computing", "C", "人工智能", "Artificial Intelligence"),
        ("WI", "Web Intelligence", "C", "人工智能", "Artificial Intelligence"),
        ("TIIS", "ACM Transactions on Interactive Intelligent Systems", "C", "人工智能", "Artificial Intelligence"),
        ("TELO", "ACM Transactions on Evolutionary Learning and Optimization", "C", "人工智能", "Artificial Intelligence"),
        ("JATS", "ACM Journal on Autonomous Transportation Systems", "C", "人工智能", "Artificial Intelligence"),
        ("TOCHI", "ACM Transactions on Computer-Human Interaction", "A", "人机交互与普适计算", "HCI & Pervasive Computing"),
        ("IJHCS", "International Journal of Human-Computer Studies", "A", "人机交互与普适计算", "HCI & Pervasive Computing"),
        ("CSCW J", "Computer Supported Cooperative Work", "B", "人机交互与普适计算", "HCI & Pervasive Computing"),
        ("HCI", "Human-Computer Interaction", "B", "人机交互与普适计算", "HCI & Pervasive Computing"),
        ("THMS", "IEEE Transactions on Human-Machine Systems", "B", "人机交互与普适计算", "HCI & Pervasive Computing"),
        ("IWC", "Interacting with Computers", "B", "人机交互与普适计算", "HCI & Pervasive Computing"),
        ("IJHCI", "International Journal of Human-Computer Interaction", "B", "人机交互与普适计算", "HCI & Pervasive Computing"),
        ("UMUAI", "User Modeling and User-Adapted Interaction", "B", "人机交互与普适计算", "HCI & Pervasive Computing"),
        ("TSMC", "IEEE Transactions on Systems, Man,and Cybernetics: Systems", "B", "人机交互与普适计算", "HCI & Pervasive Computing"),
        ("CCF TPCI", "CCF Transactions on Pervasive Computing and Interaction", "B", "人机交互与普适计算", "HCI & Pervasive Computing"),
        ("BIT", "Behaviour & Information Technology", "C", "人机交互与普适计算", "HCI & Pervasive Computing"),
        ("PUC", "Personal and Ubiquitous Computing", "C", "人机交互与普适计算", "HCI & Pervasive Computing"),
        ("PMC", "Pervasive and Mobile Computing", "C", "人机交互与普适计算", "HCI & Pervasive Computing"),
        ("PACMHCI", "Proceedings of the ACM on Human-Computer Interaction", "C", "人机交互与普适计算", "HCI & Pervasive Computing"),
        ("THRI", "ACM Transactions on Human-Robot Interaction", "C", "人机交互与普适计算", "HCI & Pervasive Computing"),
        ("JACM", "Journal of the ACM", "A", "交叉/综合/新兴", "Cross-disciplinary"),
        ("Proc. IEEE", "Proceedings of the IEEE", "A", "交叉/综合/新兴", "Cross-disciplinary"),
        ("SCIS", "Science China Information Sciences", "A", "交叉/综合/新兴", "Cross-disciplinary"),
        ("Bioinformatics", "Bioinformatics", "A", "交叉/综合/新兴", "Cross-disciplinary"),
        ("Briefings in Bioinformatics", "Briefings in Bioinformatics", "B", "交叉/综合/新兴", "Cross-disciplinary"),
        ("Cognition", "Cognition", "B", "交叉/综合/新兴", "Cross-disciplinary"),
        ("TASE J", "IEEE Transactions on Automation Science and Engineering", "B", "交叉/综合/新兴", "Cross-disciplinary"),
        ("TGARS", "IEEE Transactions on Geoscience and Remote Sensing", "B", "交叉/综合/新兴", "Cross-disciplinary"),
        ("TITS", "IEEE Transactions on Intelligent Transportation Systems", "B", "交叉/综合/新兴", "Cross-disciplinary"),
        ("TMI", "IEEE Transactions on Medical Imaging", "B", "交叉/综合/新兴", "Cross-disciplinary"),
        ("TR", "IEEE Transactions on Robotics", "B", "交叉/综合/新兴", "Cross-disciplinary"),
        ("TCBB", "IEEE/ACM Transactions on Computational Biology and Bioinformatics", "B", "交叉/综合/新兴", "Cross-disciplinary"),
        ("JCST", "Journal of Computer Science and Technology", "B", "交叉/综合/新兴", "Cross-disciplinary"),
        ("JAMIA", "Journal of the American Medical Informatics Association", "B", "交叉/综合/新兴", "Cross-disciplinary"),
        ("PLOS Computational Biology", "PLOS Computational Biology", "B", "交叉/综合/新兴", "Cross-disciplinary"),
        ("The Computer Journal", "The Computer Journal", "B", "交叉/综合/新兴", "Cross-disciplinary"),
        ("WWW J", "World Wide Web", "B", "交叉/综合/新兴", "Cross-disciplinary"),
        ("FCS", "Frontiers of Computer Science", "B", "交叉/综合/新兴", "Cross-disciplinary"),
        ("BCRA", "Blockchain: Research and Applications", "B", "交叉/综合/新兴", "Cross-disciplinary"),
        ("BMC Bioinformatics", "BMC Bioinformatics", "C", "交叉/综合/新兴", "Cross-disciplinary"),
        ("Cybernetics and Systems", "Cybernetics and Systems", "C", "交叉/综合/新兴", "Cross-disciplinary"),
        ("GRSL", "IEEE Geoscience and Remote Sensing Letters", "C", "交叉/综合/新兴", "Cross-disciplinary"),
        ("JBHI", "IEEE Journal of Biomedical and Health Informatics", "C", "交叉/综合/新兴", "Cross-disciplinary"),
        ("TBD", "IEEE Transactions on Big Data", "C", "交叉/综合/新兴", "Cross-disciplinary"),
        ("IET Intelligent Transport Systems", "IET Intelligent Transport Systems", "C", "交叉/综合/新兴", "Cross-disciplinary"),
        ("JBI", "Journal of Biomedical Informatics", "C", "交叉/综合/新兴", "Cross-disciplinary"),
        ("Medical Image Analysis", "Medical Image Analysis", "C", "交叉/综合/新兴", "Cross-disciplinary"),
        ("TII", "IEEE Transactions on Industrial Informatics", "C", "交叉/综合/新兴", "Cross-disciplinary"),
        ("TCPS", "ACM Transactions on Cyber-Physical Systems", "C", "交叉/综合/新兴", "Cross-disciplinary"),
        ("TOCE", "ACM Transactions on Computing Education", "C", "交叉/综合/新兴", "Cross-disciplinary"),
        ("FITEE", "Frontiers of Information Technology & Electronic Engineering", "C", "交叉/综合/新兴", "Cross-disciplinary"),
        ("TCSS", "IEEE Transactions on Computational Social Systems", "C", "交叉/综合/新兴", "Cross-disciplinary"),
        ("TRel", "IEEE Transactions on Reliability", "C", "交叉/综合/新兴", "Cross-disciplinary"),
        ("HEALTH", "ACM Transactions on Computing for Healthcare", "C", "交叉/综合/新兴", "Cross-disciplinary"),
        ("ACM DLT", "ACM Distributed Ledger Technologies: Research and Practice", "C", "交叉/综合/新兴", "Cross-disciplinary"),
        ("IACR TCHES", "IACR Transactions on Cryptographic Hardware and Embedded Systems", "C", "交叉/综合/新兴", "Cross-disciplinary"),
        ("IACR ToSC", "IACR Transactions on Symmetric Cryptology", "C", "交叉/综合/新兴", "Cross-disciplinary"),
    ]

    for row in raw:
        name = str(row[1]).strip() if row[1] else ""
        if not name: continue

        # Fix typos
        name = name.replace("Journal of Biomedical Informatics Informatics", "Journal of Biomedical Informatics")
        if name == "Intelligence Expert Systems":
            name = "Expert Systems with Applications"
        if name == "Engineering Applications of Artificial":
            name = "Engineering Applications of Artificial Intelligence"
        if "Information and Software Technology Cambridge" in name:
            name = "Information and Software Technology"

        match, score = match_entry(name, official_journals)
        if match:
            abbr, rank, cat_zh, cat_en = match
            entry_id = generate_id(abbr)
            if entry_id in seen_ids:
                entry_id = f"{entry_id}-{slugify(cat_en)}"
            seen_ids.add(entry_id)

            results.append(OrderedDict([
                ("id", entry_id),
                ("abbreviation", abbr),
                ("full_name", name),
                ("type", "journal"),
                ("sub_type", None),
                ("ccf_rank", rank),
                ("category_zh", cat_zh),
                ("category_en", cat_en),
                ("aliases", []),
            ]))

    rank_order = {"A": 0, "B": 1, "C": 2}
    results.sort(key=lambda j: (rank_order[j["ccf_rank"]], j["abbreviation"]))
    print(f"Journals: {len(results)} matched")
    return results

def main():
    print("=" * 60)
    print("Building CCF Data JSON Files (v2 - full name matching)")
    print("=" * 60)

    print("\n--- Conferences ---")
    conferences = build_conferences()
    ranks = {"A": 0, "B": 0, "C": 0}
    for c in conferences: ranks[c["ccf_rank"]] += 1
    print(f"  Total: {len(conferences)} (A={ranks['A']}, B={ranks['B']}, C={ranks['C']})")
    print(f"  Official: 386 (A=58, B=132, C=196)")

    with open(DATA_DIR / "conferences.json", "w", encoding="utf-8") as f:
        json.dump(conferences, f, ensure_ascii=False, indent=2)

    print("\n--- Journals ---")
    journals = build_journals()
    ranks_j = {"A": 0, "B": 0, "C": 0}
    for j in journals: ranks_j[j["ccf_rank"]] += 1
    print(f"  Total: {len(journals)} (A={ranks_j['A']}, B={ranks_j['B']}, C={ranks_j['C']})")

    with open(DATA_DIR / "journals.json", "w", encoding="utf-8") as f:
        json.dump(journals, f, ensure_ascii=False, indent=2)

    metadata = {
        "version": "2026-7th",
        "description": "CCF Recommended International Conferences and Journals (7th Edition, March 2026)",
        "description_zh": "中国计算机学会推荐国际学术会议和期刊目录（第七版，2026年3月）",
        "conference_count": len(conferences),
        "journal_count": len(journals),
        "categories": [
            {"zh": "计算机体系结构", "en": "Computer Architecture"},
            {"zh": "计算机网络", "en": "Computer Networks"},
            {"zh": "网络与信息安全", "en": "Network & Information Security"},
            {"zh": "软件工程", "en": "Software Engineering"},
            {"zh": "数据库/数据挖掘", "en": "Databases & Data Mining"},
            {"zh": "计算机科学理论", "en": "Theory of Computer Science"},
            {"zh": "计算机图形学与多媒体", "en": "Computer Graphics & Multimedia"},
            {"zh": "人工智能", "en": "Artificial Intelligence"},
            {"zh": "人机交互与普适计算", "en": "HCI & Pervasive Computing"},
            {"zh": "交叉/综合/新兴", "en": "Cross-disciplinary"},
        ],
        "ccf_ranks": ["A", "B", "C"],
    }
    with open(DATA_DIR / "metadata.json", "w", encoding="utf-8") as f:
        json.dump(metadata, f, ensure_ascii=False, indent=2)

    print(f"\n✅ Done! Files in {DATA_DIR}")

if __name__ == "__main__":
    main()
