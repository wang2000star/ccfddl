import openpyxl, json, sys, os
from openpyxl.styles import Font, PatternFill

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
os.chdir(BASE)

HDR = ['venue_id','abbreviation','year','round','submission_deadline','abstract_deadline',
       'rebuttal_start','rebuttal_end','notification','camera_ready',
       'conference_start','conference_end','location','source_url','timezone']

# All verified data categorized into 4 groups
GROUPS = {
'2025submit-2025conf': [
['ches','CHES',2025,3,'2025-01-15','','2025-02-24','2025-02-28','2025-03-15','2025-04-14','2025-09-14','2025-09-18','Kuala Lumpur, Malaysia','https://ches.iacr.org/2025/','AoE'],
['ches','CHES',2025,4,'2025-04-15','','2025-05-26','2025-05-30','2025-06-15','2025-07-14','2025-09-14','2025-09-18','Kuala Lumpur, Malaysia','https://ches.iacr.org/2025/','AoE'],
['crypto','CRYPTO',2025,1,'2025-02-13','','2025-04-07','2025-04-12','2025-05-03','2025-06-05','2025-08-17','2025-08-21','Santa Barbara, CA, USA','https://crypto.iacr.org/2025/','AoE'],
['iccv','ICCV',2025,1,'2025-03-07','','2025-05-09','2025-05-16','2025-06-25','','2025-10-19','2025-10-23','Honolulu, HI, USA','https://iccv.thecvf.com/Conferences/2025/','AoE']],
'2025submit-2026conf': [
['ches','CHES',2026,1,'2025-07-15','','2025-08-28','2025-08-31','2025-09-15','2025-10-14','2026-10-11','2026-10-15','Antalya, Turkiye','https://ches.iacr.org/2026/','AoE'],
['ches','CHES',2026,2,'2025-10-15','','2025-11-27','2025-11-30','2025-12-15','2026-01-14','2026-10-11','2026-10-15','Antalya, Turkiye','https://ches.iacr.org/2026/','AoE'],
['fse','FSE',2026,1,'2025-09-01','','2025-10-05','2025-10-08','2025-11-01','2025-11-28','2026-03-23','2026-03-27','Singapore','https://fse.iacr.org/2026/','AoE'],
['fse','FSE',2026,2,'2025-11-23','','2026-01-02','2026-01-06','2026-01-23','2026-02-20','2026-03-23','2026-03-27','Singapore','https://fse.iacr.org/2026/','AoE'],
['usenix-security','USENIX Security',2026,1,'2025-08-26','','2025-11-06','2025-11-13','2025-12-04','2026-01-15','2026-08-12','2026-08-14','Baltimore, MD, USA','https://www.usenix.org/conference/usenixsecurity26/','AoE'],
['ndss','NDSS',2026,1,'2025-07-10','','2025-09-22','2025-09-29','2025-10-24','2025-12-05','2026-02-23','2026-02-26','San Diego, CA, USA','https://www.ndss-symposium.org/','AoE'],
['s-p','S&P',2026,1,'2025-06-05','','','','2025-09-01','','2026-05-18','2026-05-21','San Francisco, CA, USA','https://www.ieee-security.org/TC/SP2026/','AoE'],
['s-p','S&P',2026,2,'2025-11-13','','','','2026-02-01','','2026-05-18','2026-05-21','San Francisco, CA, USA','https://www.ieee-security.org/TC/SP2026/','AoE'],
['eurocrypt','EUROCRYPT',2026,1,'2025-10-02','','2025-12-08','2025-12-12','2026-01-29','2026-02-27','2026-05-10','2026-05-14','Rome, Italy','https://eurocrypt.iacr.org/2026/','AoE'],
['dsn','DSN',2026,1,'2025-12-04','','2026-02-13','2026-02-27','2026-03-19','2026-04-28','2026-06-22','2026-06-25','Charlotte, NC, USA','https://dsn2026.github.io/','AoE'],
['fc','FC',2026,1,'2025-09-20','','','','2025-11-24','2026-01-12','2026-03-02','2026-03-06','St. Kitts','https://ifca.ai/fc26/','AoE'],
['euros-p','EuroS&P',2026,1,'2025-11-20','','','','2026-02-15','','2026-07-06','2026-07-10','Lisbon, Portugal','https://www.ieee-security.org/TC/EuroSP2026/','AoE'],
['pkc','PKC',2026,1,'2025-10-24','','2025-12-16','2025-12-23','2026-02-13','2026-03-06','2026-05-25','2026-05-28','West Palm Beach, FL, USA','https://pkc.iacr.org/2026/','AoE'],
['csf','CSF',2026,1,'2025-07-24','','','','2025-09-25','','2026-07-26','2026-07-29','Lisbon, Portugal','https://www.ieee-security.org/TC/CSF2026/','AoE'],
['csf','CSF',2026,2,'2025-10-09','','','','2025-12-11','','2026-07-26','2026-07-29','Lisbon, Portugal','https://www.ieee-security.org/TC/CSF2026/','AoE'],
['www','WWW',2026,1,'2025-10-24','','2026-01-06','2026-01-13','2026-01-24','2026-02-14','2026-04-19','2026-04-24','Perth, Australia','https://www2026.thewebconf.org/','AoE'],
['iclr','ICLR',2026,1,'2025-09-24','2025-09-19','','','2026-01-25','','2026-04-23','2026-04-25','TBD','https://iclr.cc/Conferences/2026/','AoE'],
['cvpr','CVPR',2026,1,'2025-11-13','2025-11-07','2026-01-22','2026-01-29','2026-02-20','','2026-06-01','2026-06-30','TBD','https://cvpr.thecvf.com/Conferences/2026/','AoE']],
'2026submit-2026conf': [
['ches','CHES',2026,3,'2026-01-15','','2026-02-26','2026-03-01','2026-03-15','2026-04-14','2026-10-11','2026-10-15','Antalya, Turkiye','https://ches.iacr.org/2026/','AoE'],
['ches','CHES',2026,4,'2026-04-15','','2026-05-28','2026-05-31','2026-06-15','2026-07-14','2026-10-11','2026-10-15','Antalya, Turkiye','https://ches.iacr.org/2026/','AoE'],
['crypto','CRYPTO',2026,1,'2026-02-12','','2026-04-07','2026-04-13','2026-05-04','2026-06-08','2026-08-17','2026-08-20','Santa Barbara, CA, USA','https://crypto.iacr.org/2026/','AoE'],
['asiacrypt','ASIACRYPT',2026,1,'2026-05-21','','2026-07-17','2026-07-22','2026-08-14','2026-09-14','2026-12-07','2026-12-11','Hong Kong, China','https://asiacrypt.iacr.org/2026/','AoE'],
['tcc','TCC',2026,1,'2026-05-19','','','','2026-08-20','2026-09-15','2026-11-10','2026-11-13','TBD','https://tcc.iacr.org/2026/','AoE'],
['ccs','CCS',2026,1,'2026-01-14','2026-01-07','2026-03-17','2026-03-20','2026-04-09','2026-08-09','2026-11-15','2026-11-19','The Hague, Netherlands','https://www.sigsac.org/ccs/CCS2026/','AoE'],
['ccs','CCS',2026,2,'2026-04-29','2026-04-22','2026-06-29','2026-07-01','2026-07-17','2026-09-13','2026-11-15','2026-11-19','The Hague, Netherlands','https://www.sigsac.org/ccs/CCS2026/','AoE'],
['sigcomm','SIGCOMM',2026,1,'2026-02-06','','2026-04-27','2026-04-29','2026-05-11','','2026-08-17','2026-08-21','Denver, CO, USA','https://conferences.sigcomm.org/sigcomm/2026/','AoE'],
['csf','CSF',2026,3,'2026-01-29','','','','2026-04-01','','2026-07-26','2026-07-29','Lisbon, Portugal','https://www.ieee-security.org/TC/CSF2026/','AoE'],
['usenix-security','USENIX Security',2026,2,'2026-02-05','','2026-04-16','2026-04-23','2026-05-14','2026-06-11','2026-08-12','2026-08-14','Baltimore, MD, USA','https://www.usenix.org/conference/usenixsecurity26/','AoE'],
['icml','ICML',2026,1,'2026-01-28','2026-01-23','','','2026-04-30','2026-05-28','2026-07-07','2026-07-09','Seoul, South Korea','https://icml.cc/Conferences/2026/','AoE'],
['neurips','NeurIPS',2026,1,'2026-05-06','2026-05-04','','','2026-09-24','','2026-12-11','2026-12-13','Sydney, Australia','https://neurips.cc/Conferences/2026/','AoE'],
['acl','ACL',2026,1,'2026-01-05','','','','2026-04-04','2026-04-19','2026-07-02','2026-07-07','San Diego, CA, USA','https://2026.aclweb.org/','AoE']],
'2026submit-2027conf': [
['ches','CHES',2027,1,'2026-07-15','','2026-08-23','2026-08-28','2026-09-15','2026-10-14','2027-09-06','2027-09-09','Cancun, Mexico','https://ches.iacr.org/2027/','AoE'],
['ches','CHES',2027,2,'2026-10-15','','2026-11-23','2026-11-28','2026-12-15','2027-01-14','2027-09-06','2027-09-09','Cancun, Mexico','https://ches.iacr.org/2027/','AoE'],
['ches','CHES',2027,3,'2027-01-15','','2027-02-23','2027-02-28','2027-03-15','2027-04-14','2027-09-06','2027-09-09','Cancun, Mexico','https://ches.iacr.org/2027/','AoE'],
['ches','CHES',2027,4,'2027-04-15','','2027-05-23','2027-05-28','2027-06-15','2027-07-14','2027-09-06','2027-09-09','Cancun, Mexico','https://ches.iacr.org/2027/','AoE'],
['fse','FSE',2027,1,'2026-03-01','','2026-04-05','2026-04-08','2026-05-01','2026-05-28','2027-05-24','2027-05-28','Maastricht, Netherlands','https://fse.iacr.org/2027/','AoE'],
['fse','FSE',2027,2,'2026-06-01','','2026-07-05','2026-07-08','2026-08-01','2026-08-28','2027-05-24','2027-05-28','Maastricht, Netherlands','https://fse.iacr.org/2027/','AoE'],
['fse','FSE',2027,3,'2026-09-01','','2026-10-05','2026-10-08','2026-11-01','2026-11-28','2027-05-24','2027-05-28','Maastricht, Netherlands','https://fse.iacr.org/2027/','AoE'],
['fse','FSE',2027,4,'2026-12-01','','2027-01-05','2027-01-08','2027-02-01','2027-02-28','2027-05-24','2027-05-28','Maastricht, Netherlands','https://fse.iacr.org/2027/','AoE'],
['eurocrypt','EUROCRYPT',2027,1,'2026-09-17','','2026-11-30','2026-12-05','2027-01-18','2027-02-08','2027-04-11','2027-04-15','Eindhoven, Netherlands','https://eurocrypt.iacr.org/2027/','AoE'],
['s-p','S&P',2027,1,'2026-06-11','','2026-08-20','2026-08-25','2026-09-11','2026-10-16','2027-05-01','2027-05-04','Montreal, Canada','https://sp2027.ieee-security.org/','AoE'],
['s-p','S&P',2027,2,'2026-11-17','','2027-02-11','2027-02-16','2027-03-05','2027-04-08','2027-05-01','2027-05-04','Montreal, Canada','https://sp2027.ieee-security.org/','AoE'],
['usenix-security','USENIX Security',2027,1,'2026-08-25','','2026-11-05','2026-11-12','2026-12-03','2027-01-14','2027-08-11','2027-08-13','Denver, CO, USA','https://www.usenix.org/conference/usenixsecurity27/','AoE'],
['usenix-security','USENIX Security',2027,2,'2027-01-26','','2027-04-08','2027-04-15','2027-05-06','2027-06-03','2027-08-11','2027-08-13','Denver, CO, USA','https://www.usenix.org/conference/usenixsecurity27/','AoE'],
['ndss','NDSS',2027,1,'2026-05-06','','2026-07-15','2026-07-17','2026-07-29','2026-09-30','2027-03-22','2027-03-26','Seoul, South Korea','https://www.ndss-symposium.org/ndss2027/submissions/call-for-papers/','AoE'],
['ndss','NDSS',2027,2,'2026-08-19','','','','2026-11-04','2027-01-06','2027-03-22','2027-03-26','Seoul, South Korea','https://www.ndss-symposium.org/ndss2027/submissions/call-for-papers/','AoE']]}

def write_excel():
    wb = openpyxl.Workbook()
    hb = Font(bold=True, color='FFFFFF', size=10)
    hf = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    wb.remove(wb.active)
    total = 0
    for sn in ['2025submit-2025conf','2025submit-2026conf','2026submit-2026conf','2026submit-2027conf']:
        rows = GROUPS[sn]
        ws = wb.create_sheet(title=sn)
        for ci, col in enumerate(HDR, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font = hb; cell.fill = hf
        for ri, rd in enumerate(rows, 2):
            for ci in range(len(rd)):
                ws.cell(row=ri, column=ci+1, value=rd[ci])
        ws.freeze_panes = 'A2'
        total += len(rows)
        print(f'{sn}: {len(rows)} entries')
    wb.save('timeline_data.xlsx')
    print(f'Total: {total}')

def write_json():
    """Write combined JSON from the 4 groups"""
    entries = []
    for sn, rows in GROUPS.items():
        for rd in rows:
            entry = {HDR[i]: rd[i] for i in range(len(HDR)) if rd[i] != ''}
            entries.append(entry)
    entries.sort(key=lambda e: (e.get('submission_deadline',''), e['venue_id']))

    import os
    d = os.path.join('data', 'timelines')
    os.makedirs(d, exist_ok=True)

    with open(os.path.join(d, 'all.json'), 'w', encoding='utf-8') as f:
        json.dump(entries, f, ensure_ascii=False, indent=2)

    years = sorted(set(e['year'] for e in entries))
    for y in years:
        ye = [e for e in entries if e['year'] == y]
        with open(os.path.join(d, f'{y}.json'), 'w', encoding='utf-8') as f:
            json.dump(ye, f, ensure_ascii=False, indent=2)

    print(f'JSON: {len(entries)} entries, years {years}')
    for y in years:
        cnt = len([e for e in entries if e['year'] == y])
        print(f'  {y}: {cnt} entries')

if __name__ == '__main__':
    write_excel()
    write_json()
