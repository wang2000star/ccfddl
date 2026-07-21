import openpyxl, json
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent
DATA_DIR = BASE / "data" / "timelines"

# Load verified data from all.json
with open(DATA_DIR / "all.json", 'r', encoding='utf-8') as f:
    all_data = json.load(f)

# Categorize by submit year → conference year
groups = {
    '2025submit-2025conf': [],
    '2025submit-2026conf': [],
    '2026submit-2026conf': [],
    '2026submit-2027conf': [],
}

for e in all_data:
    sd = e.get('submission_deadline', '')
    cy = e.get('year', '')
    if not sd or not cy:
        continue
    sy = sd[:4]
    cy_str = str(cy)
    key = sy + 'submit-' + cy_str + 'conf'
    if key in groups:
        groups[key].append(e)

# Columns - fixed order
cols = ['venue_id','abbreviation','year','round','submission_deadline','abstract_deadline',
        'rebuttal_start','rebuttal_end','notification','camera_ready',
        'conference_start','conference_end','location','source_url','timezone']

# Stats fields are also stored but as part of stats object in JSON
stats_fields = ['submissions','accepted','acceptance_rate']

wb = openpyxl.Workbook()
hb = Font(bold=True, color='FFFFFF', size=10)
hf = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
wb.remove(wb.active)

# Remove old all.json to force clean rebuild
(DATA_DIR / "all.json").unlink(missing_ok=True)

for sheet_name in sorted(groups.keys()):
    entries = groups[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    # Headers
    header_cols = cols + stats_fields
    for ci, col in enumerate(header_cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = hb
        cell.fill = hf
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for ri, entry in enumerate(entries, 2):
        stats = entry.get('stats', {})
        for ci, col in enumerate(header_cols, 1):
            if col in stats:
                val = stats[col]
            else:
                val = entry.get(col, '')
            ws.cell(row=ri, column=ci, value=val if val is not None else '')

    ws.freeze_panes = 'A2'
    print(f'{sheet_name}: {len(entries)} entries')

wb.save(BASE / "timeline_data.xlsx")
print(f'\nSaved timeline_data.xlsx with {len(wb.sheetnames)} sheets')

# Recreate all.json from Excel to ensure consistency
import sys
sys.path.insert(0, str(BASE / 'scripts'))
exec(open(BASE / 'scripts' / 'sync_from_excel.py').read())
excel_to_json()
