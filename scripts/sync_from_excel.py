#!/usr/bin/env python3
"""Read timeline data from Excel and generate JSON files."""
import json, os
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    exit(1)

BASE = Path(__file__).resolve().parent.parent
EXCEL_PATH = BASE / "timeline_data.xlsx"
DATA_DIR = BASE / "data" / "timelines"
DATA_DIR.mkdir(parents=True, exist_ok=True)

def excel_to_json():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # Read header
    headers = [cell.value for cell in ws[1]]
    print(f"Columns: {headers}")

    # Read data rows
    entries = []
    stats_keys = ['submissions', 'accepted', 'acceptance_rate']
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:  # skip empty rows
            continue

        entry = {}
        stats = {}
        for i, key in enumerate(headers):
            if key is None:
                continue
            val = row[i] if i < len(row) else None
            if val is None or val == '':
                continue
            if key in stats_keys:
                if isinstance(val, (int, float)):
                    stats[key] = val
            elif key == 'round':
                entry[key] = int(val)
            elif key == 'year':
                entry[key] = int(val)
            else:
                entry[key] = str(val).strip()

        if stats:
            entry['stats'] = stats

        required = ['venue_id', 'abbreviation', 'year']
        if all(k in entry for k in required):
            entries.append(entry)
        else:
            print(f"  SKIP (missing required field): {entry.get('venue_id','?')}")

    # Sort
    entries.sort(key=lambda e: (e['venue_id'], e['year'], e.get('round', 1)))

    # Save all.json
    all_path = DATA_DIR / "all.json"
    with open(all_path, 'w', encoding='utf-8') as f:
        json.dump(entries, f, ensure_ascii=False, indent=2)

    # Save per-year files
    years = sorted(set(e['year'] for e in entries))
    for year in years:
        year_entries = [e for e in entries if e['year'] == year]
        with open(DATA_DIR / f"{year}.json", 'w', encoding='utf-8') as f:
            json.dump(year_entries, f, ensure_ascii=False, indent=2)

    # Print summary
    print(f"\n[OK] Synced {len(entries)} timeline entries:")
    for y in years:
        count = len([e for e in entries if e['year'] == y])
        venues = len(set(e['venue_id'] for e in entries if e['year'] == y))
        print(f"  {y}: {count} entries, {venues} venues")
    print(f"\nFiles written: {all_path}")
    for y in years:
        print(f"  {DATA_DIR / f'{y}.json'}")

if __name__ == '__main__':
    excel_to_json()
